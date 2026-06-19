import os
import pickle
import re
import statistics
import zipfile
from collections import Counter
import argparse
import importlib.util
import random
import time
import difflib
import glob
import csv
import sys
import unicodedata
import hashlib
import threading
import queue
import traceback
import inspect

import openpyxl

try:
    import sentencepiece as spm
    spm_import_error = None
except Exception as e:
    spm = None
    spm_import_error = str(e)


_HANGUL_RE = re.compile(r"[\uAC00-\uD7A3]")
_HANZI_RE = re.compile(r"[\u4E00-\u9FFF]")
_ASCII_WORD_DIGIT_RE = re.compile(r"[A-Za-z][A-Za-z0-9/_\-\.\+]*|\d+(?:\.\d+)?%?")
_PIECES_RE = re.compile(r"[A-Za-z][A-Za-z0-9/_\-\.\+]*|\d+(?:\.\d+)?%?|[\uAC00-\uD7A3]+|[\u4E00-\u9FFF]+")


def _configure_console_utf8():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def _timestamped_score_output_path(prefix: str, base_dir: str | None = None) -> str:
    out_dir = base_dir or r"D:\PythonProject\evaluate models performance"
    stamp = time.strftime("%Y%m%d_%H%M%S")
    return os.path.join(out_dir, f"{prefix}_{stamp}.xlsx")


def _timestamped_vocab_diagnose_output_path(base_dir: str | None = None) -> str:
    out_dir = base_dir or r"D:\PythonProject\evaluate models performance"
    stamp = time.strftime("%Y%m%d_%H%M%S")
    return os.path.join(out_dir, f"vocab_diagnose_{stamp}.xlsx")


def _timestamped_clean_output_path(input_xlsx: str | None = None, base_dir: str | None = None) -> str:
    out_dir = base_dir or r"D:\PythonProject\Cleand Corpus"
    stamp = time.strftime("%Y%m%d_%H%M%S")
    src = os.path.basename(str(input_xlsx).strip()) if input_xlsx else "cleaned.xlsx"
    base, ext = os.path.splitext(src)
    if not ext:
        ext = ".xlsx"
    return os.path.join(out_dir, f"Cleaned_{base}_{stamp}{ext}")


def _load_keywords_file(path: str):
    path = os.path.normpath(str(path).strip().strip('"').strip("'"))
    keywords = []
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            keywords.append(line)
    uniq = []
    seen = set()
    for k in keywords:
        k2 = k.lower()
        if k2 in seen:
            continue
        seen.add(k2)
        uniq.append(k)
    
    if not uniq:
        return None
        
    import re
    # Escape keywords and join with OR
    pattern = "|".join(re.escape(k.lower()) for k in uniq)
    try:
        combined_re = re.compile(pattern)
    except Exception:
        combined_re = None
        
    return {"list": uniq, "re": combined_re}


def _has_hangul(text: str) -> bool:
    return bool(text) and bool(re.search(r"[\uAC00-\uD7A3]", text))


def _has_hanzi(text: str) -> bool:
    return bool(text) and bool(re.search(r"[\u4E00-\u9FFF]", text))


def _is_english_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = text.strip()
    if not s:
        return False
    if _has_hangul(s) or _has_hanzi(s):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9/_\-\.\+\s]*", s))


def _is_number_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = re.sub(r"\s+", "", text)
    if not s:
        return False
    return bool(re.fullmatch(r"[\d\.,:%\+\-~±]+", s))


def _is_symbol_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = re.sub(r"\s+", "", text)
    if not s:
        return False
    for ch in s:
        if ch.isalnum():
            return False
        cat = unicodedata.category(ch)
        if not (cat and (cat[0] == "P" or cat in ("Sm", "Sc"))):
            return False
    return True


def _pieces_for_single_check(text: str):
    if not isinstance(text, str) or not text:
        return []
    return [p for p in _PIECES_RE.findall(text) if p]


def _ascii_words_digits(text: str):
    if not isinstance(text, str):
        return []
    return _ASCII_WORD_DIGIT_RE.findall(text)


def _hangul_ratio(text: str) -> float:
    if not isinstance(text, str) or not text:
        return 0.0
    h = len(_HANGUL_RE.findall(text))
    return h / max(1, len(text))


def _hanzi_ratio(text: str) -> float:
    if not isinstance(text, str) or not text:
        return 0.0
    h = len(_HANZI_RE.findall(text))
    return h / max(1, len(text))


def _row_hash(ko: str, zh: str) -> bytes:
    s = (ko or "") + "\n" + (zh or "")
    return hashlib.blake2b(s.encode("utf-8", errors="ignore"), digest_size=8).digest()


def score_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str | None,
    max_ko_len: int = 128,
    max_zh_len: int = 96,
):
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    wb_values = None
    wb = None
    try:
        print(f"[score] loading values (read-only): {input_xlsx} ...")
        wb_values = openpyxl.load_workbook(input_xlsx, data_only=True, read_only=True)
        print(f"[score] loading workbook (writeable): {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=False, read_only=False, keep_vba=False)
        print("[score] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        ws_values = wb_values[ws.title] if ws.title in wb_values.sheetnames else wb_values.active
        print("[score] workbooks loaded")

        def get_header(sheet):
            return [c.value for c in sheet[1]]

        header = get_header(ws)
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), 2)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답"), 4)
        ko_len_col = find_col(("字符数(ko)", "字符数（ko）", "char_ko", "ko_len"), 5)
        zh_len_col = find_col(("字符数(zh)", "字符数（zh）", "char_zh", "zh_len"), 6)

        score_cols = [
            "len_ratio",
            "over_ko_len",
            "over_zh_len",
            "keep_score",
            "has_hanja",
            "is_duplicate_pair",
            "quality_score",
            "quality_reasons",
        ]

        existing = {str(v).strip(): i + 1 for i, v in enumerate(header) if isinstance(v, str)}
        score_col_idx = {}
        next_col = ws.max_column + 1
        for c in score_cols:
            if c in existing:
                score_col_idx[c] = existing[c]
            else:
                ws.cell(row=1, column=next_col).value = c
                score_col_idx[c] = next_col
                next_col += 1

        def get_text(cell_value):
            return "" if cell_value is None else str(cell_value).strip()

        min_col = min(ko_col, zh_fix_col, ko_len_col, zh_len_col)
        max_col = max(ko_col, zh_fix_col, ko_len_col, zh_len_col)
        off_ko = ko_col - min_col
        off_zh = zh_fix_col - min_col
        off_ko_len = ko_len_col - min_col
        off_zh_len = zh_len_col - min_col

        dup_counter = Counter()
        total_rows = max(0, ws.max_row - 1)
        print("[score] counting duplicates...")
        empty_streak = 0
        rows_iter = ws_values.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True)
        for i, row in enumerate(rows_iter, start=1):
            if i % 2000 == 0:
                print(f"[score][dup] {i} rows scanned...")
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break
                continue
                
            ko_s = get_text(row[off_ko] if off_ko < len(row) else None)
            zh_s = get_text(row[off_zh] if off_zh < len(row) else None)
            
            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
            
            empty_streak = 0
            if ko_s and zh_s:
                dup_counter[_row_hash(ko_s, zh_s)] += 1

        sum_score = 0.0
        cnt_score = 0
        over_ko = 0
        over_zh = 0
        bad_ratio = 0
        keep_cnt = 0
        keep_sum = 0.0

        print("[score] scoring rows...")
        empty_streak = 0
        rows_iter = ws_values.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True)
        for i, row in enumerate(rows_iter, start=1):
            if i % 2000 == 0:
                print(f"[score] {i} rows scored...")
            r = i + 1
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break
                continue

            ko_s = get_text(row[off_ko] if off_ko < len(row) else None)
            zh_s = get_text(row[off_zh] if off_zh < len(row) else None)
            
            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
                
            empty_streak = 0

            ko_len_raw = row[off_ko_len] if off_ko_len < len(row) else None
            zh_len_raw = row[off_zh_len] if off_zh_len < len(row) else None
            try:
                ko_len = int(ko_len_raw) if ko_len_raw is not None else len(ko_s)
            except Exception:
                ko_len = len(ko_s)
            try:
                zh_len = int(zh_len_raw) if zh_len_raw is not None else len(zh_s)
            except Exception:
                zh_len = len(zh_s)
            lr = (zh_len / max(1, ko_len)) if ko_len else 0.0

            over_ko_len = int(ko_len > int(max_ko_len))
            over_zh_len = int(zh_len > int(max_zh_len))
            over_ko += over_ko_len
            over_zh += over_zh_len

            is_dup = int((ko_s and zh_s) and dup_counter[_row_hash(ko_s, zh_s)] > 1)
            has_hanja = int(_has_hanzi(ko_s))

            keep = None
            src_tokens = set(_ascii_words_digits(ko_s))
            if src_tokens:
                keep_cnt += 1
                out_tokens = set(_ascii_words_digits(zh_s))
                keep = sum(1 for t in src_tokens if t in out_tokens) / max(1, len(src_tokens))
                keep_sum += keep

            score = 100.0
            reasons = []

            if not ko_s or not zh_s:
                score = 0.0
                reasons.append("empty")
            else:
                hr = _hangul_ratio(ko_s)
                zr = _hanzi_ratio(zh_s)
                if hr < 0.2:
                    score -= 20
                    reasons.append("low_hangul_ratio")
                if zr < 0.2:
                    score -= 20
                    reasons.append("low_hanzi_ratio")
                if ko_s == zh_s:
                    score -= 30
                    reasons.append("ko_eq_zh")
                if lr < 0.25 or lr > 3.5:
                    score -= 15
                    reasons.append("len_ratio_out_of_range")
                    bad_ratio += 1
                if keep is not None:
                    score -= 15 * (1.0 - keep)
                    if keep < 1.0:
                        reasons.append("ascii_digit_mismatch")
                if is_dup:
                    score -= 10
                    reasons.append("duplicate_pair")
                if has_hanja:
                    score -= 15
                    reasons.append("ko_has_hanja")

                pieces = _pieces_for_single_check(ko_s)
                if len(pieces) == 1:
                    score -= 8
                    reasons.append("single_like")

                if over_ko_len:
                    score -= 20
                    reasons.append("over_max_ko_len")
                if over_zh_len:
                    score -= 20
                    reasons.append("over_max_zh_len")

            if score < 0:
                score = 0.0
            if score > 100:
                score = 100.0

            sum_score += score
            cnt_score += 1

            ws.cell(row=r, column=score_col_idx["len_ratio"]).value = round(float(lr), 2)
            ws.cell(row=r, column=score_col_idx["over_ko_len"]).value = over_ko_len
            ws.cell(row=r, column=score_col_idx["over_zh_len"]).value = over_zh_len
            ws.cell(row=r, column=score_col_idx["keep_score"]).value = "" if keep is None else round(float(keep), 2)
            ws.cell(row=r, column=score_col_idx["has_hanja"]).value = has_hanja
            ws.cell(row=r, column=score_col_idx["is_duplicate_pair"]).value = is_dup
            ws.cell(row=r, column=score_col_idx["quality_score"]).value = round(float(score), 2)
            ws.cell(row=r, column=score_col_idx["quality_reasons"]).value = "|".join(reasons)

        avg_score = sum_score / max(1, cnt_score)
        keep_avg = (keep_sum / max(1, keep_cnt)) if keep_cnt else 1.0

        try:
            print(f"[score] saving to: {output_xlsx} ...")
            wb.save(output_xlsx)
            print("[score] saved")
        except PermissionError:
            print("\nCannot write file (maybe opened by Excel):", output_xlsx)
            print("Please close Excel and retry.")
            return

        print("\n[score_corpus_xlsx]")
        print("input:", input_xlsx)
        print("sheet:", ws.title)
        print("output:", output_xlsx)
        print("rows:", cnt_score)
        print("avg_quality_score:", round(avg_score, 2))
        print("over_max_ko_len_rate:", round(over_ko / max(1, cnt_score), 6), "max_ko_len:", int(max_ko_len))
        print("over_max_zh_len_rate:", round(over_zh / max(1, cnt_score), 6), "max_zh_len:", int(max_zh_len))
        print("len_ratio_out_of_range_rate:", round(bad_ratio / max(1, cnt_score), 6), "range:", "0.25~3.5")
        print("avg_keep_score(only_when_src_has_ascii_or_digits):", round(keep_avg, 2), "samples:", keep_cnt)
    finally:
        try:
            wb_values.close()
        except Exception:
            pass
        wb.close()


def tag_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str | None,
    keywords_path: str,
):
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    kw_data = _load_keywords_file(keywords_path)
    keywords = kw_data["list"] if kw_data else []
    combined_re = kw_data["re"] if kw_data else None

    tag_cols = [
        "tag",
        "is_single",
        "is_english_only",
        "is_number_only",
        "is_symbol_only",
        "has_hangul",
        "has_hanzi",
        "has_hanja",
        "has_domain_kw",
        "domain_kw_hits",
    ]
    wb = None
    try:
        print(f"[tag] loading workbook: {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=False, read_only=False, keep_vba=False)
        print("[tag] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        header = [c.value for c in ws[1]]
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), 2)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답"), 4)

        existing = {str(v).strip(): i + 1 for i, v in enumerate(header) if isinstance(v, str)}
        tag_col_idx = {}
        next_col = ws.max_column + 1
        for c in tag_cols:
            if c in existing:
                tag_col_idx[c] = existing[c]
            else:
                ws.cell(row=1, column=next_col).value = c
                tag_col_idx[c] = next_col
                next_col += 1

        def get_text(v):
            return "" if v is None else str(v).strip()

        tag_counter = Counter()
        total_rows = max(0, ws.max_row - 1)
        print(f"[tag] max_row reported by excel: {ws.max_row}")
        
        # Determine max_col for iter_rows
        max_col_to_read = max(ko_col, zh_fix_col)
        
        # Use iter_rows to avoid cell-by-cell overhead and stop on empty rows
        rows_iter = ws.iter_rows(min_row=2, min_col=1, max_col=max_col_to_read, values_only=True)
        empty_streak = 0
        for i, row in enumerate(rows_iter, start=1):
            r = i + 1
            if i % 2000 == 0:
                print(f"[tag] {i} rows processed...")
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break # Stop if 100 consecutive rows are truly empty
                continue
            
            ko_s = get_text(row[ko_col-1] if ko_col-1 < len(row) else None)
            zh_s = get_text(row[zh_fix_col-1] if zh_fix_col-1 < len(row) else None)

            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
            
            empty_streak = 0 # Reset streak if we find data

            has_h = _has_hangul(ko_s)
            has_z = _has_hanzi(zh_s)
            has_hanja = _has_hanzi(ko_s)
            is_en = _is_english_only(ko_s)
            is_num = _is_number_only(ko_s)
            is_sym = _is_symbol_only(ko_s)
            pieces = _pieces_for_single_check(ko_s)
            is_single = int(len(pieces) == 1)

            hits = []
            ko_low = ko_s.lower()
            if combined_re and combined_re.search(ko_low):
                for k in keywords:
                    if k.lower() in ko_low:
                        hits.append(k)
                        if len(hits) >= 8:
                            break
            has_kw = int(len(hits) > 0)

            if not ko_s or not zh_s:
                tag = "empty"
            elif is_en and not has_h:
                tag = "english"
            elif is_sym:
                tag = "symbol"
            elif is_num:
                tag = "number"
            elif is_single:
                tag = "single_word"
            elif has_kw:
                tag = "automation"
            elif has_h:
                tag = "life"
            else:
                tag = "other"

            tag_counter[tag] += 1
            ws.cell(row=r, column=tag_col_idx["tag"]).value = tag
            ws.cell(row=r, column=tag_col_idx["is_single"]).value = is_single
            ws.cell(row=r, column=tag_col_idx["is_english_only"]).value = int(is_en)
            ws.cell(row=r, column=tag_col_idx["is_number_only"]).value = int(is_num)
            ws.cell(row=r, column=tag_col_idx["is_symbol_only"]).value = int(is_sym)
            ws.cell(row=r, column=tag_col_idx["has_hangul"]).value = int(has_h)
            ws.cell(row=r, column=tag_col_idx["has_hanzi"]).value = int(has_z)
            ws.cell(row=r, column=tag_col_idx["has_hanja"]).value = int(has_hanja)
            ws.cell(row=r, column=tag_col_idx["has_domain_kw"]).value = has_kw
            ws.cell(row=r, column=tag_col_idx["domain_kw_hits"]).value = "|".join(hits)

        try:
            print(f"[tag] saving to: {output_xlsx} ...")
            wb.save(output_xlsx)
            print("[tag] saved")
        except PermissionError:
            print("\nCannot write file (maybe opened by Excel):", output_xlsx)
            print("Please close Excel and retry.")
            return

        print("\n[tag_corpus_xlsx]")
        print("input:", input_xlsx)
        print("sheet:", ws.title)
        print("output:", output_xlsx)
        print("keywords_file:", os.path.normpath(keywords_path))
        print("\n[tag counts]")
        for k, v in tag_counter.most_common():
            print(f"{k}: {v}")
    finally:
        wb.close()


def summarize_corpus_xlsx(input_xlsx: str, sheet_name: str | None = None) -> None:
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    if not os.path.exists(input_xlsx):
        raise RuntimeError(f"找不到文件: {input_xlsx}")

    wb = None
    try:
        print(f"[summary] loading workbook: {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=True, read_only=True)
        print("[summary] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        header = [c.value for c in ws[1]]
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), None)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답", "zh", "中文"), None)
        tag_col = find_col(("tag",), None)
        has_kw_col = find_col(("has_domain_kw",), None)
        has_hanja_col = find_col(("has_hanja",), None)

        score_cols = {
            "len_ratio": find_col(("len_ratio",), None),
            "over_ko_len": find_col(("over_ko_len",), None),
            "over_zh_len": find_col(("over_zh_len",), None),
            "keep_score": find_col(("keep_score",), None),
            "is_duplicate_pair": find_col(("is_duplicate_pair",), None),
            "quality_score": find_col(("quality_score",), None),
        }

        def get_text(v):
            return "" if v is None else str(v).strip()

        def to_float(v):
            try:
                if v is None or v == "":
                    return None
                return float(v)
            except Exception:
                return None

        def to_int01(v):
            try:
                return int(v) != 0
            except Exception:
                return False

        tag_counter = Counter()
        total_ko = 0
        total_zh = 0
        qs = []
        keep = []
        lr = []
        over_ko = 0
        over_zh = 0
        dup = 0
        has_kw = 0
        has_hanja_cnt = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            ko = get_text(row[ko_col - 1]) if ko_col and ko_col - 1 < len(row) else ""
            if not ko:
                continue
            total_ko += 1

            zh = get_text(row[zh_fix_col - 1]) if zh_fix_col and zh_fix_col - 1 < len(row) else ""
            if zh:
                total_zh += 1

            if tag_col and tag_col - 1 < len(row):
                tag = get_text(row[tag_col - 1])
                if tag:
                    tag_counter[tag] += 1

            c = score_cols.get("quality_score")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    qs.append(v)

            c = score_cols.get("keep_score")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    keep.append(v)

            c = score_cols.get("len_ratio")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    lr.append(v)

            c = score_cols.get("over_ko_len")
            if c and c - 1 < len(row):
                over_ko += int(to_int01(row[c - 1]))

            c = score_cols.get("over_zh_len")
            if c and c - 1 < len(row):
                over_zh += int(to_int01(row[c - 1]))

            c = score_cols.get("is_duplicate_pair")
            if c and c - 1 < len(row):
                dup += int(to_int01(row[c - 1]))

            if has_kw_col and has_kw_col - 1 < len(row):
                has_kw += int(to_int01(row[has_kw_col - 1]))

            if has_hanja_col and has_hanja_col - 1 < len(row):
                has_hanja_cnt += int(to_int01(row[has_hanja_col - 1]))

        def pct(x, den):
            return round((x / max(1, den)) * 100.0, 2)

        def qstat(a: list[float]):
            if not a:
                return None
            a = sorted(a)
            n = len(a)

            def q(p):
                k = (n - 1) * p
                f = int(math.floor(k))
                c = int(math.ceil(k))
                if f == c:
                    return a[f]
                return a[f] * (c - k) + a[c] * (k - f)

            return {
                "avg": round(statistics.mean(a), 2),
                "p50": round(q(0.5), 2),
                "p10": round(q(0.1), 2),
                "p90": round(q(0.9), 2),
                "min": round(a[0], 2),
                "max": round(a[-1], 2),
            }

        print("\n[Corpus Summary]")
        print("file:", input_xlsx)
        print("sheet:", ws.title)
        print("rows_with_ko:", total_ko)
        print("zh_fill_rate_%:", pct(total_zh, total_ko))
        print("tag_counts_top10:", tag_counter.most_common(10))
        print("quality_score:", qstat(qs))
        print("quality_ge_80_%:", pct(sum(1 for v in qs if v >= 80), len(qs)))
        print("quality_ge_90_%:", pct(sum(1 for v in qs if v >= 90), len(qs)))
        print("keep_score:", qstat(keep))
        print("len_ratio:", qstat(lr))
        print("len_ratio_out_of_range_%:", pct(sum(1 for v in lr if v < 0.25 or v > 3.5), len(lr)))
        print("over_ko_len_%:", pct(over_ko, total_ko))
        print("over_zh_len_%:", pct(over_zh, total_ko))
        print("duplicate_pair_%:", pct(dup, total_ko))
        print("has_domain_kw_%:", pct(has_kw, total_ko))
        print("has_hanja_%:", pct(has_hanja_cnt, total_ko))
    finally:
        wb.close()


def clean_text(sentence: str) -> str:
    if not isinstance(sentence, str):
        return ""
    sentence = re.sub(r"[^\w\s\uAC00-\uD7A3\u4e00-\u9fa5]", "", sentence)
    return sentence.strip()


def try_build_tokenizers():
    ko_name = "split"
    zh_name = "char"

    def tok_ko(s: str):
        return s.split()

    def tok_zh(s: str):
        return list(s)

    try:
        from konlpy.tag import Okt

        okt = Okt()

        def tok_ko(s: str):
            return okt.morphs(s)

        ko_name = "Okt"
    except Exception as e:
        ko_name = f"split (Okt unavailable: {e})"

    try:
        import jieba

        def tok_zh(s: str):
            return jieba.lcut(s)

        zh_name = "jieba"
    except Exception as e:
        zh_name = f"char (jieba unavailable: {e})"

    return tok_ko, tok_zh, ko_name, zh_name


def summarize_lengths(name: str, seqs):
    lens = [len(x) for x in seqs]
    lens_sorted = sorted(lens)

    def pct(p: int):
        idx = int(round((p / 100) * (len(lens_sorted) - 1)))
        return lens_sorted[idx]

    print(f"\n[{name}] length (tokens)")
    print("min/mean/median/max:", min(lens), sum(lens) / len(lens), statistics.median(lens), max(lens))
    for p in (50, 75, 90, 95, 99):
        print(f"p{p}:", pct(p))
    for th in (50, 80, 100, 150, 200):
        over = sum(1 for L in lens if L > th)
        print(f"> {th}: {over} ({over/len(lens)*100:.2f}%)")


def _length_stats(seqs):
    if not seqs:
        return {
            "count": 0,
            "min": 0,
            "mean": 0.0,
            "median": 0.0,
            "max": 0,
            "p50": 0,
            "p75": 0,
            "p90": 0,
            "p95": 0,
            "p99": 0,
            "gt_50": 0,
            "gt_80": 0,
            "gt_100": 0,
            "gt_150": 0,
            "gt_200": 0,
        }
    lens = [len(x) for x in seqs]
    lens_sorted = sorted(lens)

    def pct(p: int):
        idx = int(round((p / 100) * (len(lens_sorted) - 1)))
        return lens_sorted[idx]

    return {
        "count": len(lens),
        "min": min(lens),
        "mean": sum(lens) / len(lens),
        "median": statistics.median(lens),
        "max": max(lens),
        "p50": pct(50),
        "p75": pct(75),
        "p90": pct(90),
        "p95": pct(95),
        "p99": pct(99),
        "gt_50": sum(1 for L in lens if L > 50),
        "gt_80": sum(1 for L in lens if L > 80),
        "gt_100": sum(1 for L in lens if L > 100),
        "gt_150": sum(1 for L in lens if L > 150),
        "gt_200": sum(1 for L in lens if L > 200),
    }


def _vocab_contains(vocab, token: str) -> bool:
    try:
        if hasattr(vocab, "stoi") and isinstance(vocab.stoi, dict):
            return token in vocab.stoi
        if hasattr(vocab, "token_to_idx") and isinstance(vocab.token_to_idx, dict):
            return token in vocab.token_to_idx
        return token in vocab
    except Exception:
        return False


def _unk_stats_from_token_seqs(tok_seqs, vocab=None, topn: int = 30, explicit_unk_tokens: set[str] | None = None):
    total = 0
    unk_count = 0
    sent_with_unk = 0
    unk_counter = Counter()
    explicit_unk_tokens = explicit_unk_tokens or set()

    for toks in tok_seqs:
        has_unk = False
        for t in toks:
            total += 1
            is_unk = (t in explicit_unk_tokens) or (vocab is not None and (not _vocab_contains(vocab, t)))
            if is_unk:
                unk_count += 1
                unk_counter[t] += 1
                has_unk = True
        if has_unk:
            sent_with_unk += 1

    return {
        "total_tokens": total,
        "unk_tokens": unk_count,
        "unk_rate": ((unk_count / total) if total else 0.0),
        "sent_with_unk": sent_with_unk,
        "sent_with_unk_rate": ((sent_with_unk / len(tok_seqs)) if tok_seqs else 0.0),
        "top_unk": unk_counter.most_common(topn),
    }


def _print_unk_stats(name: str, stats: dict):
    print(f"\n[{name}] vocab coverage")
    print("total tokens:", stats["total_tokens"])
    print("unk tokens:", stats["unk_tokens"], f"({stats['unk_rate']*100:.2f}%)")
    print("sentences with any unk:", stats["sent_with_unk"], f"({stats['sent_with_unk_rate']*100:.2f}%)")
    print("top unk tokens:")
    for t, c in stats["top_unk"]:
        print(f"  {t!r}: {c}")


def unk_report(name: str, tok_seqs, vocab, topn: int = 30):
    total = 0
    unk_count = 0
    sent_with_unk = 0
    unk_counter = Counter()

    for toks in tok_seqs:
        has_unk = False
        for t in toks:
            total += 1
            if t not in vocab:
                unk_count += 1
                unk_counter[t] += 1
                has_unk = True
        if has_unk:
            sent_with_unk += 1

    print(f"\n[{name}] vocab coverage")
    print("total tokens:", total)
    print("unk tokens:", unk_count, f"({(unk_count/total*100) if total else 0:.2f}%)")
    print("sentences with any unk:", sent_with_unk, f"({sent_with_unk/len(tok_seqs)*100:.2f}%)")
    print("top unk tokens:")
    for t, c in unk_counter.most_common(topn):
        print(f"  {t!r}: {c}")


def _read_parallel_corpus_xlsx(xlsx_path: str, max_rows: int | None = None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    ko_sents, zh_sents = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 4:
            continue
        ko, zh = row[1], row[3]
        if ko and zh:
            ko_sents.append(str(ko))
            zh_sents.append(str(zh))
        if max_rows and len(ko_sents) >= max_rows:
            break
    wb.close()
    return ko_sents, zh_sents


def _save_vocab_diagnose_xlsx(output_xlsx: str, meta_rows: list[tuple[str, str]], summary_rows: list[dict], ko_top_unk, zh_top_unk):
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = "Meta"
    ws_meta.append(["key", "value"])
    for k, v in meta_rows:
        ws_meta.append([k, v])

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["lang", "metric", "value"])
    for row in summary_rows:
        ws_sum.append([row["lang"], row["metric"], row["value"]])

    ws_ko = wb.create_sheet("KO_Top_UNK")
    ws_ko.append(["token", "count"])
    for token, count in ko_top_unk:
        ws_ko.append([token, count])

    ws_zh = wb.create_sheet("ZH_Top_UNK")
    ws_zh.append(["token", "count"])
    for token, count in zh_top_unk:
        ws_zh.append([token, count])

    wb.save(output_xlsx)


def vocab_diagnose(root: str, corpus_xlsx: str, model_dir: str, output_xlsx: str | None = None):
    corpus_path = os.path.join(root, corpus_xlsx)
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))

    model_kind = "v3_vocab"
    model_path = None
    ko_vocab_path = None
    zh_vocab_path = None
    ko_spm_path = None
    zh_spm_path = None

    try:
        _md, model_path_bpe, ko_spm_bpe, zh_spm_bpe = _resolve_bpe_model_paths(model_dir)
        if os.path.exists(ko_spm_bpe) and os.path.exists(zh_spm_bpe):
            model_kind = "spm"
            ko_spm_path = ko_spm_bpe
            zh_spm_path = zh_spm_bpe
            model_path = model_path_bpe if os.path.exists(model_path_bpe) else None
    except Exception:
        pass

    try:
        _md, model_path_v4, _ckpt_path, ko_spm_v4, zh_spm_v4 = _resolve_v4_transformer_paths(model_dir)
        if model_path_v4 and os.path.exists(ko_spm_v4) and os.path.exists(zh_spm_v4):
            model_kind = "spm"
            model_path = model_path_v4
            ko_spm_path = ko_spm_v4
            zh_spm_path = zh_spm_v4
    except Exception:
        pass

    if model_kind == "v3_vocab":
        try:
            model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
        except Exception:
            print("Failed to locate model/vocab under model_dir:", model_dir)
            print("Tip: V3(pkl) 需要 ko_vocab.pkl/zh_vocab.pkl；V3.2.1(BPE) 需要 spm_ko_v3_2_1.model/spm_zh_v3_2_1.model；V4(Transformer) 需要 spm_ko_v4.model/spm_zh_v4.model。")
            raise

    print("corpus_path:", corpus_path, "exists=", os.path.exists(corpus_path))
    print("model_kind:", model_kind)
    print("model_path:", model_path, "exists=", os.path.exists(model_path) if model_path else False)
    if model_kind == "v3_vocab":
        print("ko_vocab_path:", ko_vocab_path, "exists=", os.path.exists(ko_vocab_path))
        print("zh_vocab_path:", zh_vocab_path, "exists=", os.path.exists(zh_vocab_path))
    else:
        print("ko_spm_path:", ko_spm_path, "exists=", os.path.exists(ko_spm_path))
        print("zh_spm_path:", zh_spm_path, "exists=", os.path.exists(zh_spm_path))

    ko_sents, zh_sents = _read_parallel_corpus_xlsx(corpus_path, max_rows=None)
    print("pairs read:", len(ko_sents))
    ko_sents = [clean_text(s) for s in ko_sents]
    zh_sents = [clean_text(s) for s in zh_sents]
    nonempty = [(k, z) for k, z in zip(ko_sents, zh_sents) if k and z]
    ko_sents = [k for k, _ in nonempty]
    zh_sents = [z for _, z in nonempty]
    print("pairs after clean+drop empty:", len(ko_sents))

    if model_kind == "v3_vocab":
        with open(ko_vocab_path, "rb") as f:
            ko_vocab = pickle.load(f)
        with open(zh_vocab_path, "rb") as f:
            zh_vocab = pickle.load(f)
        print("ko_vocab size:", len(ko_vocab))
        print("zh_vocab size:", len(zh_vocab))

        tok_ko, tok_zh, ko_name, zh_name = try_build_tokenizers()
        print("tokenizer ko:", ko_name)
        print("tokenizer zh:", zh_name)
        print("tokenizing...")
        ko_tok = [tok_ko(s) for s in ko_sents]
        zh_tok = [tok_zh(s) for s in zh_sents]
        ko_unk_stats = _unk_stats_from_token_seqs(ko_tok, vocab=ko_vocab)
        zh_unk_stats = _unk_stats_from_token_seqs(zh_tok, vocab=zh_vocab)
        ko_vocab_size = len(ko_vocab)
        zh_vocab_size = len(zh_vocab)
    else:
        if spm is None:
            raise RuntimeError(f"缺少 sentencepiece，无法诊断 SentencePiece 词表。错误: {spm_import_error}")
        ko_sp = spm.SentencePieceProcessor(model_file=ko_spm_path)
        zh_sp = spm.SentencePieceProcessor(model_file=zh_spm_path)
        ko_name = "SentencePiece"
        zh_name = "SentencePiece"
        print("tokenizer ko:", ko_name)
        print("tokenizer zh:", zh_name)
        print("tokenizing...")
        ko_tok = [ko_sp.encode(s, out_type=str) for s in ko_sents]
        zh_tok = [zh_sp.encode(s, out_type=str) for s in zh_sents]
        ko_unk_piece = str(ko_sp.id_to_piece(int(ko_sp.unk_id())))
        zh_unk_piece = str(zh_sp.id_to_piece(int(zh_sp.unk_id())))
        ko_unk_stats = _unk_stats_from_token_seqs(ko_tok, vocab=None, explicit_unk_tokens={ko_unk_piece})
        zh_unk_stats = _unk_stats_from_token_seqs(zh_tok, vocab=None, explicit_unk_tokens={zh_unk_piece})
        ko_vocab_size = int(ko_sp.get_piece_size())
        zh_vocab_size = int(zh_sp.get_piece_size())
        print("ko_spm vocab size:", ko_vocab_size, "unk_piece:", ko_unk_piece)
        print("zh_spm vocab size:", zh_vocab_size, "unk_piece:", zh_unk_piece)

    ko_len_stats = _length_stats(ko_tok)
    zh_len_stats = _length_stats(zh_tok)

    summarize_lengths("KO", ko_tok)
    summarize_lengths("ZH", zh_tok)
    _print_unk_stats("KO", ko_unk_stats)
    _print_unk_stats("ZH", zh_unk_stats)

    if output_xlsx:
        meta_rows = [
            ("corpus_path", corpus_path),
            ("model_dir", model_dir),
            ("model_kind", model_kind),
            ("model_path", str(model_path or "")),
            ("tokenizer_ko", ko_name),
            ("tokenizer_zh", zh_name),
            ("ko_vocab_size", str(ko_vocab_size)),
            ("zh_vocab_size", str(zh_vocab_size)),
            ("pairs_read", str(len(ko_sents))),
        ]
        if ko_vocab_path:
            meta_rows.append(("ko_vocab_path", str(ko_vocab_path)))
        if zh_vocab_path:
            meta_rows.append(("zh_vocab_path", str(zh_vocab_path)))
        if ko_spm_path:
            meta_rows.append(("ko_spm_path", str(ko_spm_path)))
        if zh_spm_path:
            meta_rows.append(("zh_spm_path", str(zh_spm_path)))

        summary_rows = []
        for lang, len_stats, unk_stats in (("KO", ko_len_stats, ko_unk_stats), ("ZH", zh_len_stats, zh_unk_stats)):
            for key in ("count", "min", "mean", "median", "max", "p50", "p75", "p90", "p95", "p99", "gt_50", "gt_80", "gt_100", "gt_150", "gt_200"):
                summary_rows.append({"lang": lang, "metric": key, "value": len_stats[key]})
            summary_rows.append({"lang": lang, "metric": "total_tokens", "value": unk_stats["total_tokens"]})
            summary_rows.append({"lang": lang, "metric": "unk_tokens", "value": unk_stats["unk_tokens"]})
            summary_rows.append({"lang": lang, "metric": "unk_rate", "value": unk_stats["unk_rate"]})
            summary_rows.append({"lang": lang, "metric": "sent_with_unk", "value": unk_stats["sent_with_unk"]})
            summary_rows.append({"lang": lang, "metric": "sent_with_unk_rate", "value": unk_stats["sent_with_unk_rate"]})

        _save_vocab_diagnose_xlsx(output_xlsx, meta_rows, summary_rows, ko_unk_stats["top_unk"], zh_unk_stats["top_unk"])
        print("\n已保存:", output_xlsx)


def _load_module_from_path(module_name: str, file_path: str):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Failed to load module: {module_name} from {file_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_module_without_gui_entry(module_name: str, file_path: str):
    import types

    with open(file_path, "r", encoding="utf-8") as f:
        source = f.read()

    lines = source.splitlines()
    cut_idx = None
    for i, line in enumerate(lines):
        s = line.strip()
        if s == "root = tk.Tk()" or s.startswith("root = tk.Tk("):
            cut_idx = i
            break

    if cut_idx is None:
        return _load_module_from_path(module_name, file_path)

    partial_source = "\n".join(lines[:cut_idx]) + "\n"
    module = types.ModuleType(module_name)
    module.__file__ = file_path
    module.__name__ = module_name
    exec(compile(partial_source, file_path, "exec"), module.__dict__)
    return module

def _infer_model_paths(model_dir: str, prefer_tag: str | None = None):
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    if not os.path.isdir(model_dir):
        raise RuntimeError(f"model_dir 不是目录: {model_dir}")

    preferred = []
    if prefer_tag:
        preferred.append(prefer_tag)
    preferred += ["v3_2_1", "v3_0_1", "v3_2", "v3_attn", "best_model"]

    candidates = glob.glob(os.path.join(model_dir, "**", "*.pth"), recursive=True)
    best_model = None
    best_score = (-1, -1.0)
    for p in candidates:
        base = os.path.basename(p).lower()
        if "best_model" not in base and not base.endswith(".pth"):
            continue
        score0 = 0
        for i, tag in enumerate(preferred[::-1]):
            if tag and tag.lower() in base:
                score0 = max(score0, i + 1)
        score = (score0, os.path.getmtime(p))
        if score > best_score:
            best_score = score
            best_model = p

    if not best_model:
        raise RuntimeError(f"在目录中找不到 .pth 模型: {model_dir}")

    ko_vocab = None
    zh_vocab = None
    ko_cands = glob.glob(os.path.join(os.path.dirname(best_model), "**", "*ko*vocab*.pkl"), recursive=True) + glob.glob(
        os.path.join(model_dir, "**", "*ko*vocab*.pkl"), recursive=True
    )
    zh_cands = glob.glob(os.path.join(os.path.dirname(best_model), "**", "*zh*vocab*.pkl"), recursive=True) + glob.glob(
        os.path.join(model_dir, "**", "*zh*vocab*.pkl"), recursive=True
    )

    def pick_vocab(cands, prefer):
        best = None
        best_sc = (-1, -1.0)
        for p in cands:
            base = os.path.basename(p).lower()
            sc0 = 0
            for i, tag in enumerate(prefer[::-1]):
                if tag and tag.lower() in base:
                    sc0 = max(sc0, i + 1)
            sc = (sc0, os.path.getmtime(p))
            if sc > best_sc:
                best_sc = sc
                best = p
        return best

    ko_vocab = pick_vocab(ko_cands, preferred)
    zh_vocab = pick_vocab(zh_cands, preferred)
    if not ko_vocab or not zh_vocab:
        raise RuntimeError(f"找到模型但没找到 vocab：model={best_model}, ko_vocab={ko_vocab}, zh_vocab={zh_vocab}")

    return best_model, ko_vocab, zh_vocab


def _read_eval_set(eval_path: str, sheet: str | None = None):
    eval_path = os.path.normpath(str(eval_path).strip().strip('"').strip("'"))
    if not os.path.exists(eval_path):
        raise RuntimeError(f"找不到 eval-set 文件: {eval_path}")

    items = []
    ext = os.path.splitext(eval_path)[1].lower()
    if ext in (".txt", ".tsv"):
        with open(eval_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if "\t" in line:
                    ko, ref = line.split("\t", 1)
                    items.append({"ko": ko.strip(), "ref": ref.strip()})
                else:
                    items.append({"ko": line, "ref": ""})
        return items

    if ext == ".csv":
        with open(eval_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return items
        header = [str(x).strip().lower() for x in rows[0]]
        ko_idx = None
        ref_idx = None
        for i, h in enumerate(header):
            if h in ("ko", "korean", "source", "src", "韩文", "韩语"):
                ko_idx = i
            if h in ("zh", "chinese", "target", "tgt", "ref", "中文"):
                ref_idx = i
        start_row = 1 if ko_idx is not None else 0
        for r in rows[start_row:]:
            if not r:
                continue
            ko = r[ko_idx] if ko_idx is not None and ko_idx < len(r) else r[0]
            ref = r[ref_idx] if ref_idx is not None and ref_idx < len(r) else ""
            ko = "" if ko is None else str(ko).strip()
            if not ko:
                continue
            items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
        return items

    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(eval_path, data_only=True, read_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            header = [str(x).strip().lower() if x is not None else "" for x in (header_row or [])]
            ko_idx = None
            ref_idx = None
            for i, h in enumerate(header):
                if h in ("ko", "korean", "source", "src", "韩文", "韩语"):
                    ko_idx = i
                if h in ("zh", "chinese", "target", "tgt", "ref", "中文"):
                    ref_idx = i
            has_header = ko_idx is not None
            if not has_header:
                if header_row:
                    ko = header_row[0] if len(header_row) > 0 else ""
                    ref = header_row[1] if len(header_row) > 1 else ""
                    ko = "" if ko is None else str(ko).strip()
                    if ko:
                        items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
            for row in rows_iter:
                if not row:
                    continue
                ko = row[ko_idx] if ko_idx is not None and ko_idx < len(row) else row[0]
                ref = row[ref_idx] if ref_idx is not None and ref_idx < len(row) else ""
                ko = "" if ko is None else str(ko).strip()
                if not ko:
                    continue
                items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
            return items
        finally:
            wb.close()

    raise RuntimeError(f"不支持的 eval-set 格式: {ext}，请用 .txt/.tsv/.csv/.xlsx")


def _repetition_bigram_ratio(text: str):
    if not isinstance(text, str):
        return 0.0, 0
    s = re.sub(r"\s+", "", text)
    if len(s) < 2:
        return 0.0, len(s)
    bigrams = [s[i : i + 2] for i in range(len(s) - 1)]
    rep_ratio = 1.0 - (len(set(bigrams)) / max(1, len(bigrams)))
    max_run = 1
    cur = 1
    for i in range(1, len(s)):
        if s[i] == s[i - 1]:
            cur += 1
            if cur > max_run:
                max_run = cur
        else:
            cur = 1
    return rep_ratio, max_run


def eval_set_word(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑评估。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --eval-set xxx.txt")
        return

    eval_items = _read_eval_set(eval_path, sheet=eval_sheet)
    if not eval_items:
        print("eval-set 为空，未评估。")
        return

    if n is not None and n > 0 and len(eval_items) > n:
        random.seed(seed)
        random.shuffle(eval_items)
        eval_items = eval_items[:n]

    mod = _load_module_from_path("rt_eval", translator_path)
    user_dict_path = os.path.join(root, "user_dict.md")
    user_dict = mod.load_user_dict(user_dict_path) if hasattr(mod, "load_user_dict") else ({}, {}, {}, {}, set())

    model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
    with open(ko_vocab_path, "rb") as f:
        ko_vocab = pickle.load(f)
    with open(zh_vocab_path, "rb") as f:
        zh_vocab = pickle.load(f)

    device = torch.device("cpu")
    input_dim = len(ko_vocab)
    output_dim = len(zh_vocab)
    enc_emb_dim = 512
    dec_emb_dim = 512
    hid_dim = 512
    n_layers = 1

    if not all(hasattr(mod, x) for x in ("Attention", "Encoder", "Decoder", "Seq2Seq")):
        raise RuntimeError(f"translator 脚本缺少模型类: {translator_path}")

    attn = mod.Attention(hid_dim)
    enc = mod.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0)
    dec = mod.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn)
    model = mod.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval()

    t0 = time.perf_counter()
    rows = []
    for it in eval_items:
        ko = it.get("ko", "")
        ref = it.get("ref", "")
        try:
            out = mod.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict, max_len=max_len, show_tokens=False)
        except Exception as e:
            print(f"[eval] 翻译异常: {type(e).__name__}: {e}", flush=True)
            print("KO:", ko, flush=True)
            out = "[ERROR]"

        keep_src = _ascii_words_digits(ko)
        keep_out = _ascii_words_digits(out)
        keep_src_set = set(keep_src)
        keep_score = sum(1 for x in keep_src_set if x in set(keep_out)) / max(1, len(keep_src_set))

        rep_ratio, max_run = _repetition_bigram_ratio(out)
        unk = "<unk>" in (out or "")
        out_has_q = "?" in (out or "")
        src_has_q = "?" in (ko or "")

        src_len = len(re.sub(r"\s+", "", str(ko)))
        out_len = len(re.sub(r"\s+", "", str(out)))
        len_ratio_src = out_len / max(1, src_len)
        ref_len = len(re.sub(r"\s+", "", str(ref))) if ref else 0
        len_ratio_ref = (out_len / max(1, ref_len)) if ref_len else None

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out": out,
                "unk": int(unk),
                "src_qmark": int(src_has_q),
                "out_qmark": int(out_has_q),
                "keep_score": keep_score,
                "rep_bigram": rep_ratio,
                "max_run": max_run,
                "src_len": src_len,
                "out_len": out_len,
                "len_ratio_src": len_ratio_src,
                "len_ratio_ref": len_ratio_ref,
            }
        )
    t1 = time.perf_counter()

    print("\n[Eval 总结]")
    print("translator:", translator_path)
    print("model:", model_path)
    print("ko_vocab:", ko_vocab_path)
    print("zh_vocab:", zh_vocab_path)
    print("samples:", len(rows))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("unk rate:", sum(r["unk"] for r in rows) / max(1, len(rows)))
    print("out '?' rate:", sum(r["out_qmark"] for r in rows) / max(1, len(rows)))
    print("avg keep(ascii/digit) score:", sum(r["keep_score"] for r in rows) / max(1, len(rows)))
    print("avg rep_bigram:", sum(r["rep_bigram"] for r in rows) / max(1, len(rows)))
    print("avg len_ratio(src):", sum(r["len_ratio_src"] for r in rows) / max(1, len(rows)))

    def collapse_flag(r):
        return r["out_len"] <= 1 or r["len_ratio_src"] < 0.25

    print("collapse rate:", sum(1 for r in rows if collapse_flag(r)) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (collapse_flag(r), r["unk"], r["rep_bigram"], -r["keep_score"]), reverse=True)
    show = min(20, len(rows_sorted))
    print(f"\n[问题样本 Top {show}]")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        if r["ref"]:
            print("REF:", r["ref"])
        print("OUT:", r["out"])
        print(
            "unk:",
            r["unk"],
            "keep:",
            round(r["keep_score"], 2),
            "rep_bigram:",
            round(r["rep_bigram"], 2),
            "max_run:",
            r["max_run"],
            "len_ratio_src:",
            round(r["len_ratio_src"], 2),
        )

    if eval_out:
        eval_out = os.path.normpath(str(eval_out).strip().strip('"').strip("'"))
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Eval"
        ws.append(
            [
                "ko",
                "ref",
                "out",
                "unk",
                "src_qmark",
                "out_qmark",
                "keep_score",
                "rep_bigram",
                "max_run",
                "src_len",
                "out_len",
                "len_ratio_src",
                "len_ratio_ref",
            ]
        )
        for r in rows:
            ws.append(
                [
                    r["ko"],
                    r["ref"],
                    r["out"],
                    r["unk"],
                    r["src_qmark"],
                    r["out_qmark"],
                    float(r["keep_score"]),
                    float(r["rep_bigram"]),
                    int(r["max_run"]),
                    int(r["src_len"]),
                    int(r["out_len"]),
                    float(r["len_ratio_src"]),
                    "" if r["len_ratio_ref"] is None else float(r["len_ratio_ref"]),
                ]
            )
        out_wb.save(eval_out)
        print("\n已保存:", eval_out)


def _unwrap_state_dict_plain(state):
    if not isinstance(state, dict):
        return state
    if any(isinstance(k, str) and k.startswith("module.") for k in state.keys()):
        return {k[7:]: v for k, v in state.items() if isinstance(k, str)}
    return state


def _infer_n_layers_from_state_dict_plain(state: dict) -> int:
    if not isinstance(state, dict):
        return 2
    max_idx = -1
    for k in state.keys():
        if not isinstance(k, str):
            continue
        m = re.match(r"^encoder\.rnn\.weight_ih_l(\d+)$", k)
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    if max_idx >= 0:
        return max_idx + 1
    for k in state.keys():
        if not isinstance(k, str):
            continue
        m = re.match(r"^decoder\.rnn\.weight_ih_l(\d+)$", k)
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    return (max_idx + 1) if max_idx >= 0 else 2


def _resolve_bpe_model_paths(model_dir: str) -> tuple[str, str, str, str]:
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    if os.path.isfile(model_dir) and model_dir.lower().endswith(".pth"):
        model_path = model_dir
        model_dir = os.path.dirname(model_path)
    else:
        model_path = os.path.join(model_dir, "best_model_v3_2_1_bpe_attn.pth")
    ko_spm_path = os.path.join(model_dir, "spm_ko_v3_2_1.model")
    zh_spm_path = os.path.join(model_dir, "spm_zh_v3_2_1.model")

    if (not os.path.exists(model_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "best_model_v3_2_1_bpe_attn.pth"), recursive=True)
        if cands:
            model_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
            model_dir = os.path.dirname(model_path)
            ko_spm_path = os.path.join(model_dir, "spm_ko_v3_2_1.model")
            zh_spm_path = os.path.join(model_dir, "spm_zh_v3_2_1.model")

    if (not os.path.exists(ko_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_ko_v3_2_1.model"), recursive=True)
        if cands:
            ko_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
    if (not os.path.exists(zh_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_zh_v3_2_1.model"), recursive=True)
        if cands:
            zh_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]

    return model_dir, model_path, ko_spm_path, zh_spm_path


def eval_set_bpe(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑评估。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --eval-set xxx.txt")
        return

    if spm is None:
        print("缺少 sentencepiece，无法在本环境跑 BPE 评估。错误:", spm_import_error)
        print("请使用装有 sentencepiece 的解释器运行。")
        return

    eval_items = _read_eval_set(eval_path, sheet=eval_sheet)
    if not eval_items:
        print("eval-set 为空，未评估。")
        return

    if n is not None and n > 0 and len(eval_items) > n:
        random.seed(seed)
        random.shuffle(eval_items)
        eval_items = eval_items[:n]

    mod = _load_module_from_path("rt_eval_bpe", translator_path)
    user_dict_path = os.path.join(root, "user_dict.md")
    user_dict = mod.load_user_dict(user_dict_path) if hasattr(mod, "load_user_dict") else ({}, {}, {}, {})

    model_dir2, model_path, ko_spm_path, zh_spm_path = _resolve_bpe_model_paths(model_dir)
    if not os.path.exists(model_path):
        raise RuntimeError(f"找不到 BPE 模型文件: {model_path}")
    if not os.path.exists(ko_spm_path) or not os.path.exists(zh_spm_path):
        raise RuntimeError(f"找不到 SentencePiece 模型: {ko_spm_path} / {zh_spm_path}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    ko_sp = spm.SentencePieceProcessor(model_file=ko_spm_path)
    zh_sp = spm.SentencePieceProcessor(model_file=zh_spm_path)

    input_dim = int(ko_sp.get_piece_size())
    output_dim = int(zh_sp.get_piece_size())
    enc_emb_dim = 512
    dec_emb_dim = 512
    hid_dim = 512

    if not all(hasattr(mod, x) for x in ("Attention", "Encoder", "Decoder", "Seq2Seq")):
        raise RuntimeError(f"translator 脚本缺少模型类: {translator_path}")

    state_obj = torch.load(model_path, map_location=device)
    if isinstance(state_obj, dict) and "state_dict" in state_obj and isinstance(state_obj["state_dict"], dict):
        state = state_obj["state_dict"]
    elif isinstance(state_obj, dict) and "model_state_dict" in state_obj and isinstance(state_obj["model_state_dict"], dict):
        state = state_obj["model_state_dict"]
    else:
        state = state_obj
    state = _unwrap_state_dict_plain(state)

    infer_fn = getattr(mod, "infer_n_layers_from_state_dict", None) or getattr(mod, "infer_n_layers_from_state_dict_plain", None)
    if callable(infer_fn):
        try:
            n_layers = int(infer_fn(state))
        except Exception:
            n_layers = _infer_n_layers_from_state_dict_plain(state if isinstance(state, dict) else {})
    else:
        n_layers = _infer_n_layers_from_state_dict_plain(state if isinstance(state, dict) else {})

    attn = mod.Attention(hid_dim)
    enc = mod.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0).to(device)
    dec = mod.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn).to(device)
    model = mod.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(state)
    model.eval()

    t0 = time.perf_counter()
    rows = []
    for it in eval_items:
        ko = it.get("ko", "")
        ref = it.get("ref", "")
        try:
            out = mod.translate_sentence(ko, model, ko_sp, zh_sp, device, user_dict, max_len=max_len)
        except Exception as e:
            print(f"[eval] 翻译异常: {type(e).__name__}: {e}", flush=True)
            print("KO:", ko, flush=True)
            out = "[ERROR]"

        keep_src = _ascii_words_digits(ko)
        keep_out = _ascii_words_digits(out)
        keep_src_set = set(keep_src)
        keep_score = sum(1 for x in keep_src_set if x in set(keep_out)) / max(1, len(keep_src_set))

        rep_ratio, max_run = _repetition_bigram_ratio(out)
        unk = "<unk>" in (out or "")
        out_has_q = ("[?]" in (out or "")) or ("?" in (out or ""))
        src_has_q = "?" in (ko or "")

        src_len = len(re.sub(r"\s+", "", str(ko)))
        out_len = len(re.sub(r"\s+", "", str(out)))
        len_ratio_src = out_len / max(1, src_len)
        ref_len = len(re.sub(r"\s+", "", str(ref))) if ref else 0
        len_ratio_ref = (out_len / max(1, ref_len)) if ref_len else None

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out": out,
                "unk": int(unk),
                "src_qmark": int(src_has_q),
                "out_qmark": int(out_has_q),
                "keep_score": keep_score,
                "rep_bigram": rep_ratio,
                "max_run": max_run,
                "src_len": src_len,
                "out_len": out_len,
                "len_ratio_src": len_ratio_src,
                "len_ratio_ref": len_ratio_ref,
            }
        )
    t1 = time.perf_counter()

    print("\n[Eval 总结]")
    print("translator:", translator_path)
    print("model:", model_path)
    print("spm_ko:", ko_spm_path)
    print("spm_zh:", zh_spm_path)
    print("samples:", len(rows))
    print("device:", str(device))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("unk rate:", sum(r["unk"] for r in rows) / max(1, len(rows)))
    print("out '?' rate:", sum(r["out_qmark"] for r in rows) / max(1, len(rows)))
    print("avg keep(ascii/digit) score:", sum(r["keep_score"] for r in rows) / max(1, len(rows)))
    print("avg rep_bigram:", sum(r["rep_bigram"] for r in rows) / max(1, len(rows)))
    print("avg len_ratio(src):", sum(r["len_ratio_src"] for r in rows) / max(1, len(rows)))

    def collapse_flag(r):
        return r["out_len"] <= 1 or r["len_ratio_src"] < 0.25

    print("collapse rate:", sum(1 for r in rows if collapse_flag(r)) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (collapse_flag(r), r["unk"], r["rep_bigram"], -r["keep_score"]), reverse=True)
    show = min(20, len(rows_sorted))
    print(f"\n[问题样本 Top {show}]")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        if r["ref"]:
            print("REF:", r["ref"])
        print("OUT:", r["out"])
        print(
            "unk:",
            r["unk"],
            "keep:",
            round(r["keep_score"], 2),
            "rep_bigram:",
            round(r["rep_bigram"], 2),
            "max_run:",
            r["max_run"],
            "len_ratio_src:",
            round(r["len_ratio_src"], 2),
        )

    if eval_out:
        eval_out = os.path.normpath(str(eval_out).strip().strip('"').strip("'"))
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Eval"
        ws.append(
            [
                "ko",
                "ref",
                "out",
                "unk",
                "src_qmark",
                "out_qmark",
                "keep_score",
                "rep_bigram",
                "max_run",
                "src_len",
                "out_len",
                "len_ratio_src",
                "len_ratio_ref",
            ]
        )
        for r in rows:
            ws.append(
                [
                    r["ko"],
                    r["ref"],
                    r["out"],
                    r["unk"],
                    r["src_qmark"],
                    r["out_qmark"],
                    float(r["keep_score"]),
                    float(r["rep_bigram"]),
                    int(r["max_run"]),
                    int(r["src_len"]),
                    int(r["out_len"]),
                    float(r["len_ratio_src"]),
                    "" if r["len_ratio_ref"] is None else float(r["len_ratio_ref"]),
                ]
            )
        out_wb.save(eval_out)
        print("\n已保存:", eval_out)


def _resolve_v4_transformer_paths(model_dir: str) -> tuple[str, str | None, str | None, str, str]:
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))

    if os.path.isfile(model_dir) and model_dir.lower().endswith(".pth"):
        model_path = model_dir
        model_dir = os.path.dirname(model_path)
        ckpt_path = os.path.join(model_dir, "best_model_v4_transformer.ckpt")
    else:
        model_path = os.path.join(model_dir, "best_model_v4_transformer.pth")
        ckpt_path = os.path.join(model_dir, "best_model_v4_transformer.ckpt")

    ko_spm_path = os.path.join(model_dir, "spm_ko_v4.model")
    zh_spm_path = os.path.join(model_dir, "spm_zh_v4.model")

    if (not os.path.exists(model_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "best_model_v4_transformer.pth"), recursive=True)
        if cands:
            model_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
            model_dir = os.path.dirname(model_path)
            ckpt_path = os.path.join(model_dir, "best_model_v4_transformer.ckpt")
            ko_spm_path = os.path.join(model_dir, "spm_ko_v4.model")
            zh_spm_path = os.path.join(model_dir, "spm_zh_v4.model")

    if (not os.path.exists(ckpt_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "best_model_v4_transformer.ckpt"), recursive=True)
        if cands:
            ckpt_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]

    if (not os.path.exists(ko_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_ko_v4.model"), recursive=True)
        if cands:
            ko_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
    if (not os.path.exists(zh_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_zh_v4.model"), recursive=True)
        if cands:
            zh_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]

    model_path = model_path if os.path.exists(model_path) else None
    ckpt_path = ckpt_path if os.path.exists(ckpt_path) else None
    return model_dir, model_path, ckpt_path, ko_spm_path, zh_spm_path


def _infer_transformer_hparams_from_state_dict_plain(state: dict) -> dict:
    d_model = None
    dim_ff = None
    enc_layers = None
    dec_layers = None

    if isinstance(state, dict):
        w = state.get("src_embedding.weight")
        if w is None:
            w = state.get("trg_embedding.weight")
        if hasattr(w, "shape") and len(w.shape) == 2:
            d_model = int(w.shape[1])

        w2 = state.get("transformer.encoder.layers.0.linear1.weight")
        if hasattr(w2, "shape") and len(w2.shape) == 2:
            dim_ff = int(w2.shape[0])

        enc_max = -1
        dec_max = -1
        for k in state.keys():
            if not isinstance(k, str):
                continue
            m = re.match(r"^transformer\.encoder\.layers\.(\d+)\.", k)
            if m:
                enc_max = max(enc_max, int(m.group(1)))
            m = re.match(r"^transformer\.decoder\.layers\.(\d+)\.", k)
            if m:
                dec_max = max(dec_max, int(m.group(1)))
        if enc_max >= 0:
            enc_layers = enc_max + 1
        if dec_max >= 0:
            dec_layers = dec_max + 1

    d_model = 256 if d_model is None else int(d_model)
    dim_ff = 1024 if dim_ff is None else int(dim_ff)
    enc_layers = 6 if enc_layers is None else int(enc_layers)
    dec_layers = 6 if dec_layers is None else int(dec_layers)

    prefer = [8, 4, 16, 2, 1]
    nhead = None
    for h in prefer:
        if d_model % int(h) == 0:
            nhead = int(h)
            break
    if nhead is None:
        nhead = 1

    return {"d_model": d_model, "dim_ff": dim_ff, "enc_layers": enc_layers, "dec_layers": dec_layers, "nhead": nhead}


def eval_set_transformer(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑评估。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --eval-set xxx.txt")
        return

    if spm is None:
        print("缺少 sentencepiece，无法在本环境跑 Transformer(BPE) 评估。错误:", spm_import_error)
        print("请使用装有 sentencepiece 的解释器运行。")
        return

    eval_items = _read_eval_set(eval_path, sheet=eval_sheet)
    if not eval_items:
        print("eval-set 为空，未评估。")
        return

    if n is not None and n > 0 and len(eval_items) > n:
        random.seed(seed)
        random.shuffle(eval_items)
        eval_items = eval_items[:n]

    mod = _load_module_from_path("rt_eval_v4_transformer", translator_path)
    user_dict_path = os.path.join(root, "user_dict.md")
    user_dict = mod.load_user_dict(user_dict_path) if hasattr(mod, "load_user_dict") else ({}, {}, {}, {})

    md, model_path, ckpt_path, ko_spm_path, zh_spm_path = _resolve_v4_transformer_paths(model_dir)
    if not model_path and not ckpt_path:
        raise RuntimeError(f"找不到 V4 Transformer 模型文件: {os.path.normpath(str(model_dir))}")
    if not os.path.exists(ko_spm_path) or not os.path.exists(zh_spm_path):
        raise RuntimeError(f"找不到 SentencePiece 模型: {ko_spm_path} / {zh_spm_path}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    ko_sp = spm.SentencePieceProcessor(model_file=ko_spm_path)
    zh_sp = spm.SentencePieceProcessor(model_file=zh_spm_path)

    if model_path:
        state_obj = torch.load(model_path, map_location="cpu")
    else:
        state_obj = torch.load(str(ckpt_path), map_location="cpu")
    if isinstance(state_obj, dict) and "model_state_dict" in state_obj and isinstance(state_obj["model_state_dict"], dict):
        state = state_obj["model_state_dict"]
    elif isinstance(state_obj, dict) and "state_dict" in state_obj and isinstance(state_obj["state_dict"], dict):
        state = state_obj["state_dict"]
    else:
        state = state_obj
    state = _unwrap_state_dict_plain(state)

    if not hasattr(mod, "TransformerModel"):
        raise RuntimeError(f"translator 脚本缺少 TransformerModel: {translator_path}")
    if not hasattr(mod, "translate_sentence"):
        raise RuntimeError(f"translator 脚本缺少 translate_sentence: {translator_path}")

    hp = _infer_transformer_hparams_from_state_dict_plain(state if isinstance(state, dict) else {})
    model = mod.TransformerModel(
        n_src_vocab=int(ko_sp.get_piece_size()),
        n_trg_vocab=int(zh_sp.get_piece_size()),
        d_model=int(hp["d_model"]),
        nhead=int(hp["nhead"]),
        num_encoder_layers=int(hp["enc_layers"]),
        num_decoder_layers=int(hp["dec_layers"]),
        dim_feedforward=int(hp["dim_ff"]),
        dropout=0.0,
    ).to(device)
    model.load_state_dict(state)
    model.eval()

    used_model_path = model_path or str(ckpt_path)
    t0 = time.perf_counter()
    rows = []
    for it in eval_items:
        ko = it.get("ko", "")
        ref = it.get("ref", "")
        try:
            out = mod.translate_sentence(ko, model, ko_sp, zh_sp, device, user_dict, max_len=max_len)
        except Exception as e:
            print(f"[eval] 翻译异常: {type(e).__name__}: {e}", flush=True)
            print("KO:", ko, flush=True)
            out = "[ERROR]"

        keep_src = _ascii_words_digits(ko)
        keep_out = _ascii_words_digits(out)
        keep_src_set = set(keep_src)
        keep_score = sum(1 for x in keep_src_set if x in set(keep_out)) / max(1, len(keep_src_set))

        rep_ratio, max_run = _repetition_bigram_ratio(out)
        unk = "<unk>" in (out or "")
        out_has_q = ("[?]" in (out or "")) or ("?" in (out or ""))
        src_has_q = "?" in (ko or "")

        src_len = len(re.sub(r"\s+", "", str(ko)))
        out_len = len(re.sub(r"\s+", "", str(out)))
        len_ratio_src = out_len / max(1, src_len)
        ref_len = len(re.sub(r"\s+", "", str(ref))) if ref else 0
        len_ratio_ref = (out_len / max(1, ref_len)) if ref_len else None

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out": out,
                "unk": int(unk),
                "src_qmark": int(src_has_q),
                "out_qmark": int(out_has_q),
                "keep_score": keep_score,
                "rep_bigram": rep_ratio,
                "max_run": max_run,
                "src_len": src_len,
                "out_len": out_len,
                "len_ratio_src": len_ratio_src,
                "len_ratio_ref": len_ratio_ref,
            }
        )
    t1 = time.perf_counter()

    print("\n[Eval 总结]")
    print("translator:", translator_path)
    print("model:", used_model_path)
    print("spm_ko:", ko_spm_path)
    print("spm_zh:", zh_spm_path)
    print("samples:", len(rows))
    print("device:", str(device))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("unk rate:", sum(r["unk"] for r in rows) / max(1, len(rows)))
    print("out '?' rate:", sum(r["out_qmark"] for r in rows) / max(1, len(rows)))
    print("avg keep(ascii/digit) score:", sum(r["keep_score"] for r in rows) / max(1, len(rows)))
    print("avg rep_bigram:", sum(r["rep_bigram"] for r in rows) / max(1, len(rows)))
    print("avg len_ratio(src):", sum(r["len_ratio_src"] for r in rows) / max(1, len(rows)))

    def collapse_flag(r):
        return r["out_len"] <= 1 or r["len_ratio_src"] < 0.25

    print("collapse rate:", sum(1 for r in rows if collapse_flag(r)) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (collapse_flag(r), r["unk"], r["rep_bigram"], -r["keep_score"]), reverse=True)
    show = min(20, len(rows_sorted))
    print(f"\n[问题样本 Top {show}]")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        if r["ref"]:
            print("REF:", r["ref"])
        print("OUT:", r["out"])
        print(
            "unk:",
            r["unk"],
            "keep:",
            round(r["keep_score"], 2),
            "rep_bigram:",
            round(r["rep_bigram"], 2),
            "max_run:",
            r["max_run"],
            "len_ratio_src:",
            round(r["len_ratio_src"], 2),
        )

    if eval_out:
        eval_out = os.path.normpath(str(eval_out).strip().strip('"').strip("'"))
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Eval"
        ws.append(
            [
                "ko",
                "ref",
                "out",
                "unk",
                "src_qmark",
                "out_qmark",
                "keep_score",
                "rep_bigram",
                "max_run",
                "src_len",
                "out_len",
                "len_ratio_src",
                "len_ratio_ref",
            ]
        )
        for r in rows:
            ws.append(
                [
                    r["ko"],
                    r["ref"],
                    r["out"],
                    r["unk"],
                    r["src_qmark"],
                    r["out_qmark"],
                    float(r["keep_score"]),
                    float(r["rep_bigram"]),
                    int(r["max_run"]),
                    int(r["src_len"]),
                    int(r["out_len"]),
                    float(r["len_ratio_src"]),
                    "" if r["len_ratio_ref"] is None else float(r["len_ratio_ref"]),
                ]
            )
        out_wb.save(eval_out)
        print("\n已保存:", eval_out)


def eval_set(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    tp = os.path.basename(str(translator_path or "")).lower()
    md = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    is_v4_transformer = (
        ("transformer" in tp)
        or os.path.exists(os.path.join(md, "best_model_v4_transformer.pth"))
        or os.path.exists(os.path.join(md, "best_model_v4_transformer.ckpt"))
        or os.path.exists(os.path.join(md, "spm_ko_v4.model"))
    )
    if is_v4_transformer:
        return eval_set_transformer(
            root=root,
            eval_path=eval_path,
            model_dir=model_dir,
            translator_path=translator_path,
            eval_out=eval_out,
            eval_sheet=eval_sheet,
            n=n,
            seed=seed,
            max_len=max_len,
        )
    is_bpe = ("bpe" in tp) or ("子词" in str(translator_path or "")) or os.path.exists(os.path.join(md, "spm_ko_v3_2_1.model"))
    if is_bpe:
        return eval_set_bpe(
            root=root,
            eval_path=eval_path,
            model_dir=model_dir,
            translator_path=translator_path,
            eval_out=eval_out,
            eval_sheet=eval_sheet,
            n=n,
            seed=seed,
            max_len=max_len,
        )
    return eval_set_word(
        root=root,
        eval_path=eval_path,
        model_dir=model_dir,
        translator_path=translator_path,
        eval_out=eval_out,
        eval_sheet=eval_sheet,
        n=n,
        seed=seed,
        max_len=max_len,
    )


def compare_rt(root: str, corpus_xlsx: str, model_dir: str, v30_path: str, v301_path: str, n: int, seed: int):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑对比。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --compare")
        return

    corpus_path = os.path.join(root, corpus_xlsx)
    model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
    user_dict_path = os.path.join(root, "user_dict.md")

    mod30 = _load_module_from_path("rt_v3_0", v30_path)
    mod301 = _load_module_from_path("rt_v3_0_1", v301_path)

    with open(ko_vocab_path, "rb") as f:
        ko_vocab = pickle.load(f)
    with open(zh_vocab_path, "rb") as f:
        zh_vocab = pickle.load(f)

    device = torch.device("cpu")
    input_dim = len(ko_vocab)
    output_dim = len(zh_vocab)
    enc_emb_dim = 512
    dec_emb_dim = 512
    hid_dim = 512
    n_layers = 1

    attn = mod301.Attention(hid_dim)
    enc = mod301.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0)
    dec = mod301.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn)
    model = mod301.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval()

    user_dict_30 = mod30.load_user_dict(user_dict_path)
    user_dict_301 = mod301.load_user_dict(user_dict_path)

    ko_sents, zh_sents = _read_parallel_corpus_xlsx(corpus_path, max_rows=None)
    pairs = [(k, z) for k, z in zip(ko_sents, zh_sents) if isinstance(k, str) and k.strip() and isinstance(z, str) and z.strip()]

    random.seed(seed)
    random.shuffle(pairs)
    pairs = pairs[: max(1, n)]

    rows = []
    t0 = time.perf_counter()
    for ko, ref in pairs:
        out30 = mod30.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict_30, show_tokens=False)
        out301 = mod301.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict_301, show_tokens=False)

        sim = difflib.SequenceMatcher(None, out30, out301).ratio()
        unk30 = "<unk>" in out30
        unk301 = "<unk>" in out301

        keep_src = _ascii_words_digits(ko)
        keep30 = _ascii_words_digits(out30)
        keep301 = _ascii_words_digits(out301)
        keep_src_set = set(keep_src)
        keep30_score = sum(1 for x in keep_src_set if x in set(keep30)) / max(1, len(keep_src_set))
        keep301_score = sum(1 for x in keep_src_set if x in set(keep301)) / max(1, len(keep_src_set))

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out30": out30,
                "out301": out301,
                "sim": sim,
                "unk30": unk30,
                "unk301": unk301,
                "keep30": keep30_score,
                "keep301": keep301_score,
            }
        )
    t1 = time.perf_counter()

    print("\n[对比总结]")
    print("samples:", len(rows))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("avg similarity(out30 vs out301):", sum(r["sim"] for r in rows) / max(1, len(rows)))
    print("unk rate v3.0:", sum(1 for r in rows if r["unk30"]) / max(1, len(rows)))
    print("unk rate v3.0.1:", sum(1 for r in rows if r["unk301"]) / max(1, len(rows)))
    print("avg ascii/digit keep v3.0:", sum(r["keep30"] for r in rows) / max(1, len(rows)))
    print("avg ascii/digit keep v3.0.1:", sum(r["keep301"] for r in rows) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (r["sim"], r["unk301"] is True), reverse=False)
    show = min(15, len(rows_sorted))
    print(f"\n[差异最大的 {show} 条样本] (sim 越小差异越大)")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        print("REF:", r["ref"])
        print("V3.0 :", r["out30"])
        print("V3.0.1:", r["out301"])
        print("sim:", round(r["sim"], 3), "unk30:", r["unk30"], "unk301:", r["unk301"], "keep30:", round(r["keep30"], 2), "keep301:", round(r["keep301"], 2))


def _load_external_translator(plugin_path: str, module_name: str):
    plugin_path = os.path.normpath(str(plugin_path).strip().strip('"').strip("'"))
    if not plugin_path or not os.path.exists(plugin_path):
        raise RuntimeError(f"翻译脚本不存在: {plugin_path}")

    mod = _load_module_without_gui_entry(module_name, plugin_path)

    class _SimpleVar:
        def __init__(self, value: str):
            self.value = value

        def get(self):
            return self.value

    # 兼容 GUI 翻译脚本：如果内部依赖 translation_direction / update_ui_status，
    # 则注入一个最小可用替身，让 diagnose_mt 可直接复用其翻译函数。
    if not hasattr(mod, "translation_direction") or not callable(getattr(getattr(mod, "translation_direction", None), "get", None)):
        setattr(mod, "translation_direction", _SimpleVar("ko2zh"))
    if not hasattr(mod, "update_ui_status"):
        setattr(mod, "update_ui_status", lambda *args, **kwargs: None)

    batch_fn = (
        getattr(mod, "translate_ko_to_zh", None)
        or getattr(mod, "translate_sentences", None)
        or getattr(mod, "get_translation_batch", None)
    )
    if callable(batch_fn):

        def _batch_translate(sentences: list[str]) -> list[str]:
            out = batch_fn(list(sentences))
            if not isinstance(out, list):
                raise RuntimeError(f"{plugin_path} 的批量翻译函数返回值不是 list")
            if len(out) != len(sentences):
                raise RuntimeError(f"{plugin_path} 的批量翻译结果数量不匹配: in={len(sentences)} out={len(out)}")
            return ["" if x is None else str(x) for x in out]

        return _batch_translate

    single_fn = (
        getattr(mod, "translate_sentence", None)
        or getattr(mod, "translate_text", None)
        or getattr(mod, "translate", None)
        or getattr(mod, "translate_ko", None)
        or getattr(mod, "get_translation", None)
    )
    if callable(single_fn):

        def _single_translate(sentences: list[str]) -> list[str]:
            out = []
            for s in sentences:
                try:
                    r = single_fn(str(s))
                except TypeError:
                    r = single_fn(text=str(s))
                out.append("" if r is None else str(r))
            return out

        return _single_translate

    raise RuntimeError(
        f"{plugin_path} 中未找到可用翻译接口。请至少实现以下之一："
        "translate_ko_to_zh(list[str]) -> list[str] / translate_sentences(list[str]) -> list[str] / "
        "get_translation_batch(list[str]) -> list[str] / translate_sentence(str) -> str / "
        "translate_text(str) -> str / get_translation(str) -> str"
    )


def _score_length_ratio_points(ratio: float, category: str) -> float:
    if category == "automatic":
        if 0.55 <= ratio <= 1.65:
            return 10.0
        if 0.40 <= ratio <= 2.10:
            return 7.0
        if 0.25 <= ratio <= 3.00:
            return 4.0
        return 0.0
    if 0.60 <= ratio <= 1.90:
        return 10.0
    if 0.45 <= ratio <= 2.40:
        return 7.0
    if 0.30 <= ratio <= 3.20:
        return 4.0
    return 0.0


def _score_structure_points(src: str, out: str) -> float:
    checks = []
    for lch, rch in [("(", ")"), ("（", "）"), ("[", "]"), ("{", "}")]:
        src_total = src.count(lch) + src.count(rch)
        if src_total <= 0:
            continue
        out_l = out.count(lch)
        out_r = out.count(rch)
        if out_l == src.count(lch) and out_r == src.count(rch):
            checks.append(1.0)
        elif out_l == out_r and (out_l + out_r) > 0:
            checks.append(0.5)
        else:
            checks.append(0.0)
    if not checks:
        return 5.0
    return round(5.0 * sum(checks) / max(1, len(checks)), 2)


def _looks_error_output(text: str) -> bool:
    s = str(text or "").strip().lower()
    if not s:
        return True
    for bad in ("[error]", "traceback", "exception", "error:", "notimplemented", "<unk>", "[?]"):
        if bad in s:
            return True
    return False


def _char_ngram_f1(pred: str, ref: str, n: int = 2) -> float:
    pred = re.sub(r"\s+", "", str(pred or ""))
    ref = re.sub(r"\s+", "", str(ref or ""))
    if not pred or not ref:
        return 0.0

    def _ngrams(text: str, k: int) -> Counter:
        if len(text) < k:
            return Counter([text]) if text else Counter()
        return Counter(text[i : i + k] for i in range(len(text) - k + 1))

    pred_ng = _ngrams(pred, n)
    ref_ng = _ngrams(ref, n)
    overlap = sum(min(c, ref_ng.get(g, 0)) for g, c in pred_ng.items())
    pred_total = sum(pred_ng.values())
    ref_total = sum(ref_ng.values())
    p = overlap / max(1, pred_total)
    r = overlap / max(1, ref_total)
    if p + r <= 0:
        return 0.0
    return (2.0 * p * r) / (p + r)


def _score_ref_length_points(ratio: float) -> float:
    if 0.85 <= ratio <= 1.20:
        return 5.0
    if 0.70 <= ratio <= 1.45:
        return 3.5
    if 0.50 <= ratio <= 1.80:
        return 2.0
    return 0.0


def _translation_effect_score(ko: str, out: str, category: str, ref: str = "") -> dict:
    ko = "" if ko is None else str(ko).strip()
    out = "" if out is None else str(out).strip()
    ref = "" if ref is None else str(ref).strip()
    notes = []

    non_empty = 15.0 if out else 0.0
    if non_empty <= 0:
        notes.append("空输出")

    hanzi_count = len(_HANZI_RE.findall(out))
    hanzi_presence = 10.0 if hanzi_count > 0 else 0.0
    if hanzi_presence <= 0:
        notes.append("缺少中文")

    hangul_left = len(_HANGUL_RE.findall(out))
    if hangul_left == 0:
        no_hangul_left = 15.0
    elif hangul_left <= 2:
        no_hangul_left = 8.0
        notes.append("残留少量韩文")
    else:
        no_hangul_left = 0.0
        notes.append("残留韩文")

    no_error_marker = 0.0 if _looks_error_output(out) else 10.0
    if no_error_marker <= 0:
        notes.append("疑似错误标记")

    src_keep = set(_ascii_words_digits(ko))
    out_keep = set(_ascii_words_digits(out))
    keep_ratio = (sum(1 for x in src_keep if x in out_keep) / max(1, len(src_keep))) if src_keep else 1.0
    ascii_digit_keep = round(15.0 * keep_ratio, 2)
    if src_keep and keep_ratio < 0.8:
        notes.append("字母/数字保留不足")

    puncts = [ch for ch in str(ko) if ch in "?!.:,;()[]{}%/-+~"]
    src_punct = Counter(puncts)
    out_punct = Counter(ch for ch in str(out) if ch in "?!.:,;()[]{}%/-+~")
    if src_punct:
        punct_keep_ratio = sum(min(v, out_punct.get(k, 0)) for k, v in src_punct.items()) / max(1, sum(src_punct.values()))
    else:
        punct_keep_ratio = 1.0
    punctuation_keep = round(10.0 * punct_keep_ratio, 2)
    if src_punct and punct_keep_ratio < 0.8:
        notes.append("标点/符号保留不足")

    src_len = len(re.sub(r"\s+", "", ko))
    out_len = len(re.sub(r"\s+", "", out))
    len_ratio = out_len / max(1, src_len)
    length_ratio_score = _score_length_ratio_points(len_ratio, category)
    if length_ratio_score < 7.0:
        notes.append("长度比例异常")

    rep_ratio, max_run = _repetition_bigram_ratio(out)
    if rep_ratio <= 0.18 and max_run <= 2:
        repetition_fluency = 10.0
    elif rep_ratio <= 0.30 and max_run <= 3:
        repetition_fluency = 6.0
        notes.append("存在重复痕迹")
    else:
        repetition_fluency = 2.0 if out else 0.0
        notes.append("重复较明显")

    structure_balance = _score_structure_points(ko, out)
    if structure_balance < 5.0:
        notes.append("括号/结构不稳定")

    if category == "automatic":
        category_fit = round(10.0 * ((keep_ratio * 0.7) + (punct_keep_ratio * 0.3)), 2)
        if category_fit < 7.0:
            notes.append("术语/单位保留一般")
    else:
        ascii_out = _ascii_words_digits(out)
        ascii_penalty = 1.0 if len(ascii_out) <= max(2, len(src_keep) + 2) else 0.6
        natural_bonus = 1.0 if (hanzi_count > 0 and rep_ratio <= 0.25) else 0.5
        category_fit = round(10.0 * ascii_penalty * natural_bonus, 2)
        if category_fit < 7.0:
            notes.append("生活句自然度一般")

    base_total = round(
        non_empty
        + hanzi_presence
        + no_hangul_left
        + no_error_marker
        + ascii_digit_keep
        + punctuation_keep
        + length_ratio_score
        + repetition_fluency
        + structure_balance
        + category_fit,
        2,
    )
    base_total = max(0.0, min(100.0, base_total))

    score_base_scaled = base_total
    score_ref_seq = 0.0
    score_ref_char_ngram = 0.0
    score_ref_length = 0.0
    ref_seq_ratio = 0.0
    ref_char_f1 = 0.0

    if ref:
        score_base_scaled = round(base_total * 0.7, 2)
        ref_seq_ratio = difflib.SequenceMatcher(None, out, ref).ratio() if out and ref else 0.0
        ref_char_f1 = _char_ngram_f1(out, ref, n=2)
        ref_len = len(re.sub(r"\s+", "", ref))
        ref_len_ratio = out_len / max(1, ref_len)
        score_ref_seq = round(15.0 * ref_seq_ratio, 2)
        score_ref_char_ngram = round(10.0 * ref_char_f1, 2)
        score_ref_length = _score_ref_length_points(ref_len_ratio)
        if ref_seq_ratio < 0.55:
            notes.append("参考答案相似度偏低")
        if ref_char_f1 < 0.55:
            notes.append("参考字符匹配偏低")
        total = round(score_base_scaled + score_ref_seq + score_ref_char_ngram + score_ref_length, 2)
    else:
        total = base_total

    total = max(0.0, min(100.0, total))

    return {
        "score_total": total,
        "score_base_scaled": score_base_scaled,
        "score_non_empty": non_empty,
        "score_hanzi_presence": hanzi_presence,
        "score_no_hangul_left": no_hangul_left,
        "score_no_error_marker": no_error_marker,
        "score_ascii_digit_keep": ascii_digit_keep,
        "score_punctuation_keep": punctuation_keep,
        "score_length_ratio": length_ratio_score,
        "score_repetition_fluency": repetition_fluency,
        "score_structure_balance": structure_balance,
        "score_category_fit": category_fit,
        "score_ref_seq": score_ref_seq,
        "score_ref_char_ngram": score_ref_char_ngram,
        "score_ref_length": score_ref_length,
        "len_ratio": round(len_ratio, 4),
        "keep_ratio": round(keep_ratio, 4),
        "punct_keep_ratio": round(punct_keep_ratio, 4),
        "rep_bigram": round(rep_ratio, 4),
        "ref_seq_ratio": round(ref_seq_ratio, 4),
        "ref_char_f1": round(ref_char_f1, 4),
        "max_run": int(max_run),
        "hangul_left": int(hangul_left),
        "notes": " | ".join(notes),
    }


def _save_single_translator_score_xlsx(output_xlsx: str, translator_name: str, rows: list[dict], elapsed_sec: float):
    if not rows:
        raise RuntimeError(f"{translator_name} 没有可输出的评分结果")

    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)

    def _grade(score: float) -> str:
        s = float(score)
        if s >= 90:
            return "优秀"
        if s >= 80:
            return "可用"
        if s >= 60:
            return "待人工复核"
        return "不建议直接使用"

    categories = []
    seen = set()
    for r in rows:
        c = str(r.get("category", ""))
        if c not in seen:
            seen.add(c)
            categories.append(c)
    if "overall" not in seen:
        categories.append("overall")

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["translator", "category", "samples", "avg_score", "excellent", "usable", "review", "reject", "elapsed_sec"])

    for category in categories:
        cat_rows = rows if category == "overall" else [r for r in rows if r.get("category") == category]
        if not cat_rows:
            continue
        grades = Counter(_grade(float(r.get("score_total", 0.0))) for r in cat_rows)
        ws_sum.append(
            [
                translator_name,
                category,
                len(cat_rows),
                round(sum(float(r.get("score_total", 0.0)) for r in cat_rows) / max(1, len(cat_rows)), 2),
                grades.get("优秀", 0),
                grades.get("可用", 0),
                grades.get("待人工复核", 0),
                grades.get("不建议直接使用", 0),
                round(float(elapsed_sec), 3),
            ]
        )

    ws_detail = wb.create_sheet("Detail")
    detail_header = [
        "category",
        "row_no",
        "ko",
        "ref",
        "out",
        "score_total",
        "grade",
        "score_base_scaled",
        "score_non_empty",
        "score_hanzi_presence",
        "score_no_hangul_left",
        "score_no_error_marker",
        "score_ascii_digit_keep",
        "score_punctuation_keep",
        "score_length_ratio",
        "score_repetition_fluency",
        "score_structure_balance",
        "score_category_fit",
        "score_ref_seq",
        "score_ref_char_ngram",
        "score_ref_length",
        "len_ratio",
        "keep_ratio",
        "punct_keep_ratio",
        "rep_bigram",
        "ref_seq_ratio",
        "ref_char_f1",
        "max_run",
        "hangul_left",
        "notes",
    ]
    ws_detail.append(detail_header)
    for r in rows:
        row = dict(r)
        row["grade"] = _grade(float(row.get("score_total", 0.0)))
        ws_detail.append([row.get(k, "") for k in detail_header])

    wb.save(output_xlsx)


def score_external_translators(
    life_eval_path: str,
    auto_eval_path: str,
    baidu_path: str,
    kimi_path: str,
    baidu_output_xlsx: str,
    kimi_output_xlsx: str,
    n: int | None = None,
    seed: int = 42,
):
    life_eval_path = os.path.normpath(str(life_eval_path).strip().strip('"').strip("'"))
    auto_eval_path = os.path.normpath(str(auto_eval_path).strip().strip('"').strip("'"))
    baidu_path = os.path.normpath(str(baidu_path).strip().strip('"').strip("'"))
    kimi_path = os.path.normpath(str(kimi_path).strip().strip('"').strip("'"))
    baidu_output_xlsx = os.path.normpath(str(baidu_output_xlsx).strip().strip('"').strip("'"))
    kimi_output_xlsx = os.path.normpath(str(kimi_output_xlsx).strip().strip('"').strip("'"))

    if not os.path.exists(life_eval_path):
        raise RuntimeError(f"生活语句文件不存在: {life_eval_path}")
    if not os.path.exists(auto_eval_path):
        raise RuntimeError(f"自动化语句文件不存在: {auto_eval_path}")
    if not os.path.exists(baidu_path):
        raise RuntimeError(f"Baidu 翻译脚本不存在: {baidu_path}")
    if not os.path.exists(kimi_path):
        raise RuntimeError(f"Kimi 翻译脚本不存在: {kimi_path}")

    baidu_translate = _load_external_translator(baidu_path, "ext_baidu_translator")
    kimi_translate = _load_external_translator(kimi_path, "ext_kimi_translator")

    baidu_rows = []
    kimi_rows = []
    datasets = [("life", life_eval_path), ("automatic", auto_eval_path)]
    t0 = time.perf_counter()

    for category, eval_path in datasets:
        items = _read_eval_set(eval_path, sheet=None)
        if not items:
            print(f"[warn] {category} 数据集为空: {eval_path}")
            continue

        if n is not None and int(n) > 0 and len(items) > int(n):
            random.seed(seed)
            random.shuffle(items)
            items = items[: int(n)]

        ko_list = [str(it.get("ko", "")).strip() for it in items]
        print(f"[external-score] {category} 样本数: {len(ko_list)}", flush=True)

        baidu_outs = baidu_translate(ko_list)
        kimi_outs = kimi_translate(ko_list)
        if len(baidu_outs) != len(ko_list) or len(kimi_outs) != len(ko_list):
            raise RuntimeError(f"{category} 翻译结果数量不匹配")

        for idx, ko in enumerate(ko_list, start=1):
            ref = "" if items[idx - 1].get("ref") is None else str(items[idx - 1].get("ref")).strip()
            baidu_out = "" if baidu_outs[idx - 1] is None else str(baidu_outs[idx - 1])
            kimi_out = "" if kimi_outs[idx - 1] is None else str(kimi_outs[idx - 1])
            baidu_score = _translation_effect_score(ko, baidu_out, category, ref=ref)
            kimi_score = _translation_effect_score(ko, kimi_out, category, ref=ref)
            baidu_row = {"category": category, "row_no": idx, "ko": ko, "ref": ref, "out": baidu_out}
            kimi_row = {"category": category, "row_no": idx, "ko": ko, "ref": ref, "out": kimi_out}
            baidu_row.update(baidu_score)
            kimi_row.update(kimi_score)
            baidu_rows.append(baidu_row)
            kimi_rows.append(kimi_row)

    t1 = time.perf_counter()
    elapsed_sec = round(t1 - t0, 3)

    _save_single_translator_score_xlsx(baidu_output_xlsx, "baidu", baidu_rows, elapsed_sec=elapsed_sec)
    _save_single_translator_score_xlsx(kimi_output_xlsx, "kimi", kimi_rows, elapsed_sec=elapsed_sec)

    print("\n[外部翻译评分总结]")
    print("life_eval_path:", life_eval_path)
    print("auto_eval_path:", auto_eval_path)
    print("baidu_path:", baidu_path)
    print("kimi_path:", kimi_path)
    print("elapsed_sec:", elapsed_sec)
    print("已保存(Baidu):", baidu_output_xlsx)
    print("已保存(Kimi):", kimi_output_xlsx)


KO_RE = re.compile(r"[\uAC00-\uD7A3]")
ZH_RE = re.compile(r"[\u4E00-\u9FFF]")
ASCII_ALPHA_RE = re.compile(r"[A-Za-z]+")
ASCII_DIGIT_RE = re.compile(r"\d+")
KEEP_PUNCT_CHARS = "?!.:,;()[]{}%/-+~\"'@#&*_="
_PUNCT_NORMALIZE_TABLE = str.maketrans(
    {
        ",": ",",
        "，": ",",
        "、": ",",
        "､": ",",
        "。": ".",
        "．": ".",
        "：": ":",
        "；": ";",
        "！": "!",
        "？": "?",
        "（": "(",
        "）": ")",
        "《": "[",
        "》": "]",
        "〈": "[",
        "〉": "]",
        "【": "[",
        "】": "]",
        "［": "[",
        "］": "]",
        "｛": "{",
        "｝": "}",
        "％": "%",
        "／": "/",
        "－": "-",
        "—": "-",
        "–": "-",
        "＋": "+",
        "～": "~",
        "·": ",",
        "・": ",",
        "“": '"',
        "”": '"',
        "‘": "'",
        "’": "'",
    }
)


def _count_regex(pattern: re.Pattern, text: str) -> int:
    if not isinstance(text, str) or not text:
        return 0
    return len(pattern.findall(text))


def _keep_ratio_from_counts(src_items, out_items) -> tuple[float, int]:
    src_counter = Counter(src_items)
    if not src_counter:
        return 1.0, 0
    out_counter = Counter(out_items)
    matched = sum(min(v, out_counter.get(k, 0)) for k, v in src_counter.items())
    total = sum(src_counter.values())
    return (matched / max(1, total)), total


def _normalized_puncts(text: str):
    if not isinstance(text, str) or not text:
        return []
    s = text.translate(_PUNCT_NORMALIZE_TABLE)
    return [ch for ch in s if ch in KEEP_PUNCT_CHARS]


def _normalized_alpha_words(text: str):
    if not isinstance(text, str) or not text:
        return []
    return [w.lower() for w in ASCII_ALPHA_RE.findall(text)]


def _normalized_digit_chunks(text: str):
    if not isinstance(text, str) or not text:
        return []
    return ASCII_DIGIT_RE.findall(text)


def _loose_dedup_key(text: str) -> str:
    if not isinstance(text, str) or not text:
        return ""
    s = unicodedata.normalize("NFKC", text).lower()
    kept = []
    for ch in s:
        if ch.isspace():
            continue
        cat = unicodedata.category(ch)
        if cat and (cat[0] in ("P", "S")):
            continue
        kept.append(ch)
    return "".join(kept)


def _loose_similarity_signatures(text: str) -> set[str]:
    if not text:
        return set()
    sigs = {f"len:{max(1, len(text) // 4)}"}
    if len(text) >= 1:
        sigs.add(f"p1:{text[:1]}")
        sigs.add(f"s1:{text[-1:]}")
    if len(text) >= 2:
        sigs.add(f"p2:{text[:2]}")
        sigs.add(f"s2:{text[-2:]}")
    if len(text) >= 4:
        mid = len(text) // 2
        sigs.add(f"m2:{text[mid - 1:mid + 1]}")
    return sigs


def _find_similar_loose_dup(
    text: str, sig_index: dict[str, list[tuple[str, int]]], threshold: float = 0.5, max_candidates: int = 200
) -> int | None:
    if not text:
        return None
    candidates = []
    seen = set()
    text_len = len(text)
    for sig in _loose_similarity_signatures(text):
        for cand_text, cand_row in sig_index.get(sig, []):
            if cand_text in seen:
                continue
            seen.add(cand_text)
            cand_len = len(cand_text)
            shorter = min(text_len, cand_len)
            longer = max(text_len, cand_len)
            if shorter / max(1, longer) < threshold:
                continue
            candidates.append((cand_text, cand_row))
            if len(candidates) >= max_candidates:
                break
        if len(candidates) >= max_candidates:
            break
    for cand_text, cand_row in candidates:
        if difflib.SequenceMatcher(None, text, cand_text).ratio() >= threshold:
            return cand_row
    return None


def _add_loose_similarity_index(text: str, row_no: int, sig_index: dict[str, list[tuple[str, int]]], max_per_sig: int = 200):
    if not text:
        return
    for sig in _loose_similarity_signatures(text):
        bucket = sig_index.setdefault(sig, [])
        if len(bucket) < max_per_sig:
            bucket.append((text, row_no))


def clean_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str | None = None,
    min_hangul_ratio: float = 0.2,
    min_hanzi_ratio: float = 0.2,
    min_punct_keep_ratio: float = 0.2,
    min_alpha_keep_ratio: float = 0.2,
    min_digit_keep_ratio: float = 0.2,
    max_ko_chars: int = 128,
    max_zh_chars: int = 96,
    prefer_zh_col: int = 3,
):
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    wb = openpyxl.load_workbook(input_xlsx, data_only=False, read_only=True)
    sheet_name = (str(sheet_name).strip() if sheet_name is not None else "")
    if not sheet_name:
        if not wb.sheetnames:
            wb.close()
            raise RuntimeError("工作簿中没有可用工作表")
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"找不到工作表: {sheet_name}，可用: {wb.sheetnames}")

    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    header = [c.value for c in header_cells] if header_cells else ["序号", "KO", "ZH", "ZH修正"]
    out_header = [*header, "delete_mark", "delete_reason"]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name
    out_ws.append(out_header)

    total = clean_rows = marked_rows = 0
    marked_by_reason = Counter()
    skipped_all_empty = 0
    empty_streak = 0
    max_empty_streak = 2000
    seen_ko = {}
    seen_zh = {}
    seen_ko_loose = {}
    seen_zh_loose = {}
    ko_loose_sig_index = {}
    zh_loose_sig_index = {}

    max_col = max(len(header), int(prefer_zh_col) + 1, 3)
    for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col, values_only=False), start=2):
        row_list = [c.value for c in row] if row else []
        if not row_list or all((v is None) or (isinstance(v, str) and (not v.strip())) for v in row_list):
            skipped_all_empty += 1
            empty_streak += 1
            if empty_streak >= max_empty_streak:
                break
            continue
        empty_streak = 0
        total += 1

        ko = row_list[1] if len(row_list) > 1 else None
        zh_pref = row_list[prefer_zh_col] if len(row_list) > prefer_zh_col else None
        zh_fallback = row_list[2] if len(row_list) > 2 else None

        ko_s = "" if ko is None else str(ko).strip()
        zh_s = "" if zh_pref is None else str(zh_pref).strip()
        if not zh_s:
            zh_s = "" if zh_fallback is None else str(zh_fallback).strip()

        reasons = []
        if not ko_s or not zh_s:
            reasons.append("empty_ko_or_zh")
        else:
            ko_loose = _loose_dedup_key(ko_s)
            zh_loose = _loose_dedup_key(zh_s)
            ko_dup_row = seen_ko.get(ko_s)
            zh_dup_row = seen_zh.get(zh_s)
            ko_near_dup_row = seen_ko_loose.get(ko_loose) if ko_loose and (ko_dup_row is None) else None
            zh_near_dup_row = seen_zh_loose.get(zh_loose) if zh_loose and (zh_dup_row is None) else None
            ko_sim_dup_row = (
                _find_similar_loose_dup(ko_loose, ko_loose_sig_index, threshold=0.5)
                if ko_loose and (ko_dup_row is None) and (ko_near_dup_row is None)
                else None
            )
            zh_sim_dup_row = (
                _find_similar_loose_dup(zh_loose, zh_loose_sig_index, threshold=0.5)
                if zh_loose and (zh_dup_row is None) and (zh_near_dup_row is None)
                else None
            )

            if ko_dup_row is not None:
                reasons.append(f"dup_ko(row={ko_dup_row})")
            else:
                seen_ko[ko_s] = excel_row_no
            if ko_near_dup_row is not None:
                reasons.append(f"near_dup_ko(row={ko_near_dup_row})")
            if ko_sim_dup_row is not None:
                reasons.append(f"similar_dup_ko(row={ko_sim_dup_row})")
            if ko_loose:
                seen_ko_loose.setdefault(ko_loose, excel_row_no)
                _add_loose_similarity_index(ko_loose, excel_row_no, ko_loose_sig_index)
            if zh_dup_row is not None:
                reasons.append(f"dup_zh(row={zh_dup_row})")
            else:
                seen_zh[zh_s] = excel_row_no
            if zh_near_dup_row is not None:
                reasons.append(f"near_dup_zh(row={zh_near_dup_row})")
            if zh_sim_dup_row is not None:
                reasons.append(f"similar_dup_zh(row={zh_sim_dup_row})")
            if zh_loose:
                seen_zh_loose.setdefault(zh_loose, excel_row_no)
                _add_loose_similarity_index(zh_loose, excel_row_no, zh_loose_sig_index)
            if ko_s == zh_s:
                reasons.append("ko_eq_zh")
            if len(ko_s) > int(max_ko_chars):
                reasons.append("ko_too_long")
            if len(zh_s) > int(max_zh_chars):
                reasons.append("zh_too_long")

            ko_h = _count_regex(KO_RE, ko_s)
            zh_h = _count_regex(ZH_RE, zh_s)

            if ko_h == 0:
                reasons.append("no_hangul")
            if zh_h == 0:
                reasons.append("no_hanzi")

            if ko_h > 0:
                ko_ratio = ko_h / max(1, len(ko_s))
                if ko_ratio < float(min_hangul_ratio):
                    reasons.append("low_hangul_ratio")
            if zh_h > 0:
                zh_ratio = zh_h / max(1, len(zh_s))
                if zh_ratio < float(min_hanzi_ratio):
                    reasons.append("low_hanzi_ratio")

            punct_keep_ratio, punct_src_total = _keep_ratio_from_counts(_normalized_puncts(ko_s), _normalized_puncts(zh_s))
            if punct_src_total > 0 and punct_keep_ratio < float(min_punct_keep_ratio):
                reasons.append("low_punct_keep_ratio")

            alpha_keep_ratio, alpha_src_total = _keep_ratio_from_counts(_normalized_alpha_words(ko_s), _normalized_alpha_words(zh_s))
            if alpha_src_total > 0 and alpha_keep_ratio < float(min_alpha_keep_ratio):
                reasons.append("low_alpha_keep_ratio")

            digit_keep_ratio, digit_src_total = _keep_ratio_from_counts(_normalized_digit_chunks(ko_s), _normalized_digit_chunks(zh_s))
            if digit_src_total > 0 and digit_keep_ratio < float(min_digit_keep_ratio):
                reasons.append("low_digit_keep_ratio")

        out_row = row_list[:]
        if len(out_row) > 1:
            out_row[1] = ko_s
        if len(out_row) > prefer_zh_col:
            out_row[prefer_zh_col] = zh_s
        delete_mark = 1 if reasons else 0
        delete_reason = "; ".join(reasons)
        out_ws.append([*out_row, delete_mark, delete_reason])
        if delete_mark:
            marked_rows += 1
            for reason in reasons:
                marked_by_reason[reason.split("(", 1)[0]] += 1
        else:
            clean_rows += 1

    wb.close()
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    out_wb.save(output_xlsx)

    print("\n[clean_corpus_xlsx]")
    print("input:", input_xlsx)
    print("sheet:", sheet_name)
    print("output:", output_xlsx)
    print("rows_total:", total)
    print("rows_clean:", clean_rows)
    print("rows_marked:", marked_rows)
    if marked_by_reason:
        print("\n[marked reasons]")
        for k, v in marked_by_reason.most_common():
            print(f"{k}: {v}")


class _QueueWriter:
    def __init__(self, q: "queue.Queue[str]"):
        self.q = q

    def write(self, s: str):
        if not s:
            return 0
        try:
            self.q.put(str(s))
        except Exception:
            pass
        return len(s)

    def flush(self):
        return


def ui_main():
    try:
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox
    except Exception as e:
        print("tkinter 不可用，无法启动 UI。错误:", e)
        return

    root = tk.Tk()
    root.title("diagnose_mt 工具")
    root.geometry("980x720")

    q: "queue.Queue[str]" = queue.Queue()
    running = {"value": False}

    def poll_log():
        try:
            while True:
                s = q.get_nowait()
                txt.insert("end", s)
                txt.see("end")
        except queue.Empty:
            pass
        root.after(100, poll_log)

    frame = ttk.Frame(root, padding=12)
    frame.pack(fill="both", expand=True)

    def clear_log():
        txt.delete("1.0", "end")

    def _norm_path(v: str):
        return os.path.normpath(str(v).strip().strip('"').strip("'"))

    def _browse_open_xlsx(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])

    def _browse_save_xlsx(title: str, initialdir: str | None = None):
        return filedialog.asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")],
        )

    def _browse_open_txt(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Text", "*.txt"), ("All", "*.*")])

    def _browse_open_py(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Python", "*.py"), ("All", "*.*")])

    def _browse_open_any(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("All", "*.*")])

    def _browse_dir(title: str):
        return filedialog.askdirectory(title=title)

    def run_in_thread(fn, args_dict: dict):
        if running["value"]:
            messagebox.showinfo("提示", "正在运行，请等待结束")
            return
        running["value"] = True
        txt.delete("1.0", "end")

        def worker():
            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = _QueueWriter(q), _QueueWriter(q)
            try:
                fn(**args_dict)
            except Exception:
                print("\n[UI] 异常：\n" + traceback.format_exc())
            finally:
                sys.stdout, sys.stderr = old_out, old_err
                running["value"] = False

        threading.Thread(target=worker, daemon=True).start()

    nb = ttk.Notebook(frame)
    nb.pack(fill="both", expand=True)

    tab1 = ttk.Frame(nb, padding=10)
    tab2 = ttk.Frame(nb, padding=10)
    tab3 = ttk.Frame(nb, padding=10)
    tab4 = ttk.Frame(nb, padding=10)
    tab5 = ttk.Frame(nb, padding=10)
    tab6 = ttk.Frame(nb, padding=10)
    nb.add(tab1, text="打标/打分")
    nb.add(tab2, text="清洗")
    nb.add(tab3, text="评估")
    nb.add(tab4, text="对比")
    nb.add(tab5, text="词表诊断")
    nb.add(tab6, text="外部评分")

    v1 = {
        "input_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "output_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "same_output": tk.BooleanVar(value=True),
        "sheet": tk.StringVar(value="Corpus"),
        "keywords_file": tk.StringVar(value=r"D:\PythonProject\corpus_sources\domain_keywords_ko.txt"),
        "max_ko_len": tk.StringVar(value="128"),
        "max_zh_len": tk.StringVar(value="96"),
        "do_tag": tk.BooleanVar(value=True),
        "do_score": tk.BooleanVar(value=True),
    }

    def v1_sync_output():
        if v1["same_output"].get():
            v1["output_xlsx"].set(v1["input_xlsx"].get())

    def v1_browse_input():
        p = _browse_open_xlsx("选择输入 xlsx")
        if p:
            v1["input_xlsx"].set(p)
            v1_sync_output()

    def v1_browse_output():
        p = _browse_save_xlsx("选择输出 xlsx")
        if p:
            v1["output_xlsx"].set(p)
            v1["same_output"].set(False)

    def v1_browse_keywords():
        p = _browse_open_txt("选择关键词文件")
        if p:
            v1["keywords_file"].set(p)

    def v1_run():
        input_xlsx = _norm_path(v1["input_xlsx"].get())
        output_xlsx = _norm_path(v1["output_xlsx"].get())
        sheet = v1["sheet"].get().strip() or None
        if not input_xlsx or not os.path.exists(input_xlsx):
            messagebox.showerror("参数错误", "输入文件不存在")
            return
        if not output_xlsx:
            messagebox.showerror("参数错误", "输出文件不能为空")
            return
        if not (v1["do_tag"].get() or v1["do_score"].get()):
            messagebox.showerror("参数错误", "至少选择一个功能（打标签/打分）")
            return
        kw = _norm_path(v1["keywords_file"].get())
        try:
            mk = int(v1["max_ko_len"].get().strip())
            mz = int(v1["max_zh_len"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "max_ko_len / max_zh_len 必须是整数")
            return

        def task():
            if v1["do_tag"].get():
                print("\n[UI] 开始：打标签")
                tag_corpus_xlsx(input_xlsx=input_xlsx, output_xlsx=output_xlsx, sheet_name=sheet, keywords_path=kw)
            if v1["do_score"].get():
                print("\n[UI] 开始：打分")
                score_corpus_xlsx(
                    input_xlsx=input_xlsx, output_xlsx=output_xlsx, sheet_name=sheet, max_ko_len=mk, max_zh_len=mz
                )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab1, text="输入文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["input_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_input).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab1, text="输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab1, text="输出=输入", variable=v1["same_output"], command=v1_sync_output).grid(
        row=r, column=1, sticky="w", padx=6
    )
    r += 1
    ttk.Label(tab1, text="Sheet").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["sheet"], width=24).grid(row=r, column=1, sticky="w", padx=6)
    r += 1
    ttk.Separator(tab1, orient="horizontal").grid(row=r, column=0, columnspan=3, sticky="we", pady=8)
    r += 1
    ttk.Checkbutton(tab1, text="打标签", variable=v1["do_tag"]).grid(row=r, column=0, sticky="w")
    ttk.Label(tab1, text="关键词文件").grid(row=r, column=1, sticky="w", padx=6)
    ttk.Entry(tab1, textvariable=v1["keywords_file"], width=70).grid(row=r, column=1, sticky="e", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_keywords).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab1, text="打分", variable=v1["do_score"]).grid(row=r, column=0, sticky="w")
    sub = ttk.Frame(tab1)
    sub.grid(row=r, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="max_ko_len").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v1["max_ko_len"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="max_zh_len").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v1["max_zh_len"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Button(tab1, text="运行本页", command=v1_run).grid(row=r, column=0, sticky="w", pady=8)
    tab1.columnconfigure(1, weight=1)

    v2 = {
        "input_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "output_xlsx": tk.StringVar(value=_timestamped_clean_output_path(r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx")),
        "same_output": tk.BooleanVar(value=True),
        "sheet": tk.StringVar(value=""),
        "min_hangul_ratio": tk.StringVar(value="0.2"),
        "min_hanzi_ratio": tk.StringVar(value="0.2"),
        "min_punct_keep_ratio": tk.StringVar(value="0.2"),
        "min_alpha_keep_ratio": tk.StringVar(value="0.2"),
        "min_digit_keep_ratio": tk.StringVar(value="0.2"),
        "max_ko_chars": tk.StringVar(value="128"),
        "max_zh_chars": tk.StringVar(value="96"),
        "prefer_zh_col": tk.StringVar(value="3"),
    }

    def v2_sync_output():
        if v2["same_output"].get():
            v2["output_xlsx"].set(_timestamped_clean_output_path(v2["input_xlsx"].get()))

    def v2_browse_input():
        p = _browse_open_xlsx("选择输入 xlsx")
        if p:
            v2["input_xlsx"].set(p)
            v2_sync_output()

    def v2_browse_output():
        p = _browse_save_xlsx("选择输出 xlsx")
        if p:
            v2["output_xlsx"].set(p)
            v2["same_output"].set(False)

    def v2_run():
        input_xlsx = _norm_path(v2["input_xlsx"].get())
        if v2["same_output"].get():
            v2_sync_output()
        output_xlsx = _norm_path(v2["output_xlsx"].get())
        sheet = v2["sheet"].get().strip() or None
        if not input_xlsx or not os.path.exists(input_xlsx):
            messagebox.showerror("参数错误", "输入文件不存在")
            return
        if not output_xlsx:
            messagebox.showerror("参数错误", "输出文件不能为空")
            return
        try:
            mh = float(v2["min_hangul_ratio"].get().strip())
            mz = float(v2["min_hanzi_ratio"].get().strip())
            mp = float(v2["min_punct_keep_ratio"].get().strip())
            ma = float(v2["min_alpha_keep_ratio"].get().strip())
            md = float(v2["min_digit_keep_ratio"].get().strip())
            mkc = int(v2["max_ko_chars"].get().strip())
            mzc = int(v2["max_zh_chars"].get().strip())
            pref = int(v2["prefer_zh_col"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "比例必须是数字，长度与 prefer_zh_col 必须是整数")
            return

        def task():
            print("\n[UI] 开始：清洗")
            clean_corpus_xlsx(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                sheet_name=sheet,
                min_hangul_ratio=mh,
                min_hanzi_ratio=mz,
                min_punct_keep_ratio=mp,
                min_alpha_keep_ratio=ma,
                min_digit_keep_ratio=md,
                max_ko_chars=mkc,
                max_zh_chars=mzc,
                prefer_zh_col=pref,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab2, text="输入文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["input_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab2, text="选择", command=v2_browse_input).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab2, text="输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab2, text="选择", command=v2_browse_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab2, text="自动生成时间戳输出", variable=v2["same_output"], command=v2_sync_output).grid(
        row=r, column=1, sticky="w", padx=6
    )
    r += 1
    ttk.Label(tab2, text="Sheet").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["sheet"], width=24).grid(row=r, column=1, sticky="w", padx=6)
    r += 1
    sub = ttk.Frame(tab2)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="min_hangul_ratio").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v2["min_hangul_ratio"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="min_hanzi_ratio").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v2["min_hanzi_ratio"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    ttk.Label(sub, text="prefer_zh_col").grid(row=0, column=4, sticky="w")
    ttk.Entry(sub, textvariable=v2["prefer_zh_col"], width=6).grid(row=0, column=5, sticky="w", padx=6)
    r += 1
    sub2 = ttk.Frame(tab2)
    sub2.grid(row=r, column=0, columnspan=3, sticky="w", pady=2)
    ttk.Label(sub2, text="min_punct_keep").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub2, textvariable=v2["min_punct_keep_ratio"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub2, text="min_alpha_keep").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub2, textvariable=v2["min_alpha_keep_ratio"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    ttk.Label(sub2, text="min_digit_keep").grid(row=0, column=4, sticky="w")
    ttk.Entry(sub2, textvariable=v2["min_digit_keep_ratio"], width=8).grid(row=0, column=5, sticky="w", padx=6)
    r += 1
    sub3 = ttk.Frame(tab2)
    sub3.grid(row=r, column=0, columnspan=3, sticky="w", pady=2)
    ttk.Label(sub3, text="max_ko_chars").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub3, textvariable=v2["max_ko_chars"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub3, text="max_zh_chars").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub3, textvariable=v2["max_zh_chars"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Button(tab2, text="运行本页", command=v2_run).grid(row=r, column=0, sticky="w", pady=8)
    tab2.columnconfigure(1, weight=1)

    v3 = {
        "eval_path": tk.StringVar(value=r"D:\PythonProject\for_live_sentences_ko.txt"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\Google_colab\V4.0_Tranformer\_1st\Epoch30_test_loss 6.5368"),
        "translator": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0.1(greedy).py"),
        "eval_out": tk.StringVar(value=r"D:\PythonProject\evaluate models performance\evaluation_result.xlsx"),
        "eval_sheet": tk.StringVar(value=""),
        "n": tk.StringVar(value="100"),
        "seed": tk.StringVar(value="42"),
        "max_len": tk.StringVar(value="80"),
    }

    def v3_browse_eval():
        p = _browse_open_any("选择评估集文件")
        if p:
            v3["eval_path"].set(p)

    def v3_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v3["model_dir"].set(p)

    def v3_browse_translator():
        p = _browse_open_py("选择翻译脚本")
        if p:
            v3["translator"].set(p)

    def v3_browse_eval_out():
        # 设置默认打开的目录
        init_dir = r"D:\PythonProject\evaluate models performance"
        if not os.path.exists(init_dir):
            try:
                os.makedirs(init_dir, exist_ok=True)
            except Exception:
                init_dir = None
        
        p = _browse_save_xlsx("选择导出评估表 xlsx", initialdir=init_dir)
        if p:
            v3["eval_out"].set(p)

    def v3_run():
        eval_path = _norm_path(v3["eval_path"].get())
        model_dir = _norm_path(v3["model_dir"].get())
        translator = _norm_path(v3["translator"].get())
        eval_out = _norm_path(v3["eval_out"].get()) if v3["eval_out"].get().strip() else None
        
        # 确保评估结果目录存在
        if eval_out:
            eval_out_dir = os.path.dirname(eval_out)
            if eval_out_dir and not os.path.exists(eval_out_dir):
                try:
                    os.makedirs(eval_out_dir, exist_ok=True)
                except Exception as e:
                    messagebox.showwarning("目录创建失败", f"无法创建目录: {eval_out_dir}\n错误: {e}")
        eval_sheet = v3["eval_sheet"].get().strip() or None
        try:
            n = int(v3["n"].get().strip())
            seed = int(v3["seed"].get().strip())
            max_len = int(v3["max_len"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "n/seed/max_len 必须是整数")
            return
        if not eval_path or not os.path.exists(eval_path):
            messagebox.showerror("参数错误", "评估集文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        if not translator or not os.path.exists(translator):
            messagebox.showerror("参数错误", "翻译脚本不存在")
            return

        def task():
            print("\n[UI] 开始：评估")
            eval_set(
                root=os.path.dirname(__file__),
                eval_path=eval_path,
                model_dir=model_dir,
                translator_path=translator,
                eval_out=eval_out,
                eval_sheet=eval_sheet,
                n=n,
                seed=seed,
                max_len=max_len,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab3, text="评估集文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["eval_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_eval).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="translator.py").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["translator"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_translator).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="eval_out(可选)").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["eval_out"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_eval_out).grid(row=r, column=2, sticky="we")
    r += 1
    sub = ttk.Frame(tab3)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="sheet(可选)").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v3["eval_sheet"], width=12).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="n").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v3["n"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    ttk.Label(sub, text="seed").grid(row=0, column=4, sticky="w")
    ttk.Entry(sub, textvariable=v3["seed"], width=8).grid(row=0, column=5, sticky="w", padx=6)
    ttk.Label(sub, text="max_len").grid(row=0, column=6, sticky="w")
    ttk.Entry(sub, textvariable=v3["max_len"], width=8).grid(row=0, column=7, sticky="w", padx=6)
    r += 1
    ttk.Button(tab3, text="运行本页", command=v3_run).grid(row=r, column=0, sticky="w", pady=8)
    tab3.columnconfigure(1, weight=1)

    v4 = {
        "corpus_xlsx": tk.StringVar(value=r"D:\PythonProject\Corpus(K2C)-2.xlsx"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\V3.0(Attention)\20260527-best Model\epoch03-4.4666"),
        "v30": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0(greedy).py"),
        "v301": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0.1(greedy).py"),
        "n": tk.StringVar(value="30"),
        "seed": tk.StringVar(value="42"),
    }

    def v4_browse_corpus():
        p = _browse_open_xlsx("选择语料 xlsx")
        if p:
            v4["corpus_xlsx"].set(p)

    def v4_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v4["model_dir"].set(p)

    def v4_browse_v30():
        p = _browse_open_py("选择 V3.0 翻译脚本")
        if p:
            v4["v30"].set(p)

    def v4_browse_v301():
        p = _browse_open_py("选择 V3.0.1 翻译脚本")
        if p:
            v4["v301"].set(p)

    def v4_run():
        corpus_path = _norm_path(v4["corpus_xlsx"].get())
        model_dir = _norm_path(v4["model_dir"].get())
        v30 = _norm_path(v4["v30"].get())
        v301 = _norm_path(v4["v301"].get())
        try:
            n = int(v4["n"].get().strip())
            seed = int(v4["seed"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "n/seed 必须是整数")
            return
        if not corpus_path or not os.path.exists(corpus_path):
            messagebox.showerror("参数错误", "语料文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        if not v30 or not os.path.exists(v30) or not v301 or not os.path.exists(v301):
            messagebox.showerror("参数错误", "翻译脚本不存在")
            return

        root_dir = os.path.dirname(corpus_path)
        corpus_name = os.path.basename(corpus_path)

        def task():
            print("\n[UI] 开始：对比")
            compare_rt(
                root=root_dir,
                corpus_xlsx=corpus_name,
                model_dir=model_dir,
                v30_path=v30,
                v301_path=v301,
                n=n,
                seed=seed,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab4, text="语料文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["corpus_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_corpus).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="V3.0 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["v30"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_v30).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="V3.0.1 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["v301"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_v301).grid(row=r, column=2, sticky="we")
    r += 1
    sub = ttk.Frame(tab4)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="n").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v4["n"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="seed").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v4["seed"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Button(tab4, text="运行本页", command=v4_run).grid(row=r, column=0, sticky="w", pady=8)
    tab4.columnconfigure(1, weight=1)

    v5 = {
        "corpus_xlsx": tk.StringVar(value=r"D:\PythonProject\Corpus(K2C)-2.xlsx"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\V3.0(Attention)\20260527-best Model\epoch03-4.4666"),
        "output_xlsx": tk.StringVar(value=_timestamped_vocab_diagnose_output_path()),
    }

    def v5_browse_corpus():
        p = _browse_open_xlsx("选择语料 xlsx")
        if p:
            v5["corpus_xlsx"].set(p)

    def v5_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v5["model_dir"].set(p)

    def v5_browse_output():
        p = _browse_save_xlsx("保存词表诊断结果", initialdir=r"D:\PythonProject\evaluate models performance")
        if p:
            v5["output_xlsx"].set(p)

    def v5_run():
        corpus_path = _norm_path(v5["corpus_xlsx"].get())
        model_dir = _norm_path(v5["model_dir"].get())
        output_xlsx = _norm_path(v5["output_xlsx"].get())
        if not corpus_path or not os.path.exists(corpus_path):
            messagebox.showerror("参数错误", "语料文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        if not output_xlsx:
            messagebox.showerror("参数错误", "请填写输出文件路径")
            return
        root_dir = os.path.dirname(corpus_path)
        corpus_name = os.path.basename(corpus_path)

        def task():
            print("\n[UI] 开始：词表诊断")
            vocab_diagnose(root=root_dir, corpus_xlsx=corpus_name, model_dir=model_dir, output_xlsx=output_xlsx)
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab5, text="语料文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab5, textvariable=v5["corpus_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab5, text="选择", command=v5_browse_corpus).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab5, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab5, textvariable=v5["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab5, text="选择", command=v5_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab5, text="输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab5, textvariable=v5["output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab5, text="选择", command=v5_browse_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Button(tab5, text="运行本页", command=v5_run).grid(row=r, column=0, sticky="w", pady=8)
    tab5.columnconfigure(1, weight=1)

    v6 = {
        "life_eval_path": tk.StringVar(value=r"D:\PythonProject\for_live_sentences_ko.standard.tsv"),
        "auto_eval_path": tk.StringVar(value=r"D:\PythonProject\for_automatic_sentences_ko.standard.tsv"),
        "baidu_path": tk.StringVar(value=r"D:\pythonproject 2\File Translation tools\（baidu）Translator-V2.12(片段翻译-保存Corpus更新）.py"),
        "kimi_path": tk.StringVar(value=r"D:\pythonproject 2\File Translation tools\（kimi）Translator-V2.14(批量请求）.py"),
        "baidu_output_xlsx": tk.StringVar(value=_timestamped_score_output_path("baidu_score")),
        "kimi_output_xlsx": tk.StringVar(value=_timestamped_score_output_path("kimi_score")),
        "n": tk.StringVar(value="0"),
        "seed": tk.StringVar(value="42"),
    }

    def v6_browse_life():
        p = _browse_open_any("选择生活语句文件")
        if p:
            v6["life_eval_path"].set(p)

    def v6_browse_auto():
        p = _browse_open_any("选择自动化语句文件")
        if p:
            v6["auto_eval_path"].set(p)

    def v6_browse_baidu():
        p = _browse_open_py("选择 Baidu 翻译脚本")
        if p:
            v6["baidu_path"].set(p)

    def v6_browse_kimi():
        p = _browse_open_py("选择 Kimi 翻译脚本")
        if p:
            v6["kimi_path"].set(p)

    def v6_browse_baidu_output():
        init_dir = r"D:\PythonProject\evaluate models performance"
        if not os.path.exists(init_dir):
            try:
                os.makedirs(init_dir, exist_ok=True)
            except Exception:
                init_dir = None
        p = _browse_save_xlsx("选择导出 Baidu 评分表 xlsx", initialdir=init_dir)
        if p:
            v6["baidu_output_xlsx"].set(p)

    def v6_browse_kimi_output():
        init_dir = r"D:\PythonProject\evaluate models performance"
        if not os.path.exists(init_dir):
            try:
                os.makedirs(init_dir, exist_ok=True)
            except Exception:
                init_dir = None
        p = _browse_save_xlsx("选择导出 Kimi 评分表 xlsx", initialdir=init_dir)
        if p:
            v6["kimi_output_xlsx"].set(p)

    def v6_run():
        life_eval_path = _norm_path(v6["life_eval_path"].get())
        auto_eval_path = _norm_path(v6["auto_eval_path"].get())
        baidu_path = _norm_path(v6["baidu_path"].get())
        kimi_path = _norm_path(v6["kimi_path"].get())
        baidu_output_xlsx = _norm_path(v6["baidu_output_xlsx"].get())
        kimi_output_xlsx = _norm_path(v6["kimi_output_xlsx"].get())
        try:
            n = int(v6["n"].get().strip())
            seed = int(v6["seed"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "n/seed 必须是整数")
            return
        if not life_eval_path or not os.path.exists(life_eval_path):
            messagebox.showerror("参数错误", "生活语句文件不存在")
            return
        if not auto_eval_path or not os.path.exists(auto_eval_path):
            messagebox.showerror("参数错误", "自动化语句文件不存在")
            return
        if not baidu_path or not os.path.exists(baidu_path):
            messagebox.showerror("参数错误", "Baidu 翻译脚本不存在")
            return
        if not kimi_path or not os.path.exists(kimi_path):
            messagebox.showerror("参数错误", "Kimi 翻译脚本不存在")
            return
        if not baidu_output_xlsx or not kimi_output_xlsx:
            messagebox.showerror("参数错误", "Baidu/Kimi 输出文件都不能为空")
            return

        def task():
            print("\n[UI] 开始：外部翻译评分")
            score_external_translators(
                life_eval_path=life_eval_path,
                auto_eval_path=auto_eval_path,
                baidu_path=baidu_path,
                kimi_path=kimi_path,
                baidu_output_xlsx=baidu_output_xlsx,
                kimi_output_xlsx=kimi_output_xlsx,
                n=(n if n > 0 else None),
                seed=seed,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab6, text="生活语句文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["life_eval_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_life).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab6, text="自动化语句文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["auto_eval_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_auto).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab6, text="Baidu 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["baidu_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_baidu).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab6, text="Kimi 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["kimi_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_kimi).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab6, text="Baidu 输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["baidu_output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_baidu_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab6, text="Kimi 输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab6, textvariable=v6["kimi_output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab6, text="选择", command=v6_browse_kimi_output).grid(row=r, column=2, sticky="we")
    r += 1
    sub = ttk.Frame(tab6)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="n(0=全部)").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v6["n"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="seed").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v6["seed"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Label(
        tab6,
        text="评分项: 空输出/中文存在/残留韩文/错误标记/字母数字保留/标点保留/长度比例/重复流畅度/结构平衡/分类适配，总分 100",
        wraplength=860,
    ).grid(row=r, column=0, columnspan=3, sticky="w", pady=(0, 6))
    r += 1
    ttk.Button(tab6, text="运行本页", command=v6_run).grid(row=r, column=0, sticky="w", pady=8)
    tab6.columnconfigure(1, weight=1)

    bottom = ttk.Frame(frame)
    bottom.pack(fill="both", expand=False, pady=(10, 0))
    btns = ttk.Frame(bottom)
    btns.pack(fill="x")
    ttk.Button(btns, text="清空输出", command=clear_log).pack(side="left")
    txt = tk.Text(bottom, height=18, wrap="word")
    txt.pack(fill="both", expand=True, pady=(8, 0))

    poll_log()
    root.mainloop()


def main():
    _configure_console_utf8()
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=r"D:\PythonProject")
    parser.add_argument("--corpus", default="Corpus(K2C)-2.xlsx")
    parser.add_argument("--model-dir", default=None)
    parser.add_argument("--vocab-diagnose-out", default=_timestamped_vocab_diagnose_output_path())
    parser.add_argument("--compare", action="store_true")
    parser.add_argument("--plugin-score", action="store_true")
    parser.add_argument("--clean-corpus", action="store_true")
    parser.add_argument("--tag-corpus", action="store_true")
    parser.add_argument("--score-corpus", action="store_true")
    parser.add_argument("--summarize-corpus", action="store_true")
    parser.add_argument("--keywords-file", default=None)
    parser.add_argument("--max-ko-len", type=int, default=128)
    parser.add_argument("--max-zh-len", type=int, default=96)
    parser.add_argument("--eval-set", default=None)
    parser.add_argument("--eval-out", default=None)
    parser.add_argument("--eval-sheet", default=None)
    parser.add_argument("--translator", default=None)
    parser.add_argument("--max-len", type=int, default=50)
    parser.add_argument("--input-xlsx", default=None)
    parser.add_argument("--output-xlsx", default=None)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--min-hangul-ratio", type=float, default=0.2)
    parser.add_argument("--min-hanzi-ratio", type=float, default=0.2)
    parser.add_argument("--min-punct-keep-ratio", type=float, default=0.2)
    parser.add_argument("--min-alpha-keep-ratio", type=float, default=0.2)
    parser.add_argument("--min-digit-keep-ratio", type=float, default=0.2)
    parser.add_argument("--clean-max-ko-chars", type=int, default=128)
    parser.add_argument("--clean-max-zh-chars", type=int, default=96)
    parser.add_argument("--n", type=int, default=30)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--v30", default=None)
    parser.add_argument("--v301", default=None)
    parser.add_argument("--life-set", default=r"D:\PythonProject\for_live_sentences_ko.standard.tsv")
    parser.add_argument("--auto-set", default=r"D:\PythonProject\for_automatic_sentences_ko.standard.tsv")
    parser.add_argument("--baidu-plugin", default=r"D:\pythonproject 2\File Translation tools\（baidu）Translator-V2.12(片段翻译-保存Corpus更新）.py")
    parser.add_argument("--kimi-plugin", default=r"D:\pythonproject 2\File Translation tools\（kimi）Translator-V2.14(批量请求）.py")
    parser.add_argument("--baidu-score-out", default=_timestamped_score_output_path("baidu_score"))
    parser.add_argument("--kimi-score-out", default=_timestamped_score_output_path("kimi_score"))
    parser.add_argument("--ui", action="store_true")
    args = parser.parse_args()

    model_dir = args.model_dir or os.path.join(args.root, "Translate Model")

    if args.ui or len(sys.argv) == 1:
        ui_main()
        return

    if args.summarize_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "NEW-CORPUS-20260604.xlsx")
        summarize_corpus_xlsx(input_xlsx=input_xlsx, sheet_name=args.sheet if args.sheet else None)
        return

    if args.score_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "Corpus(K2C)-2_20260526.xlsx")
        output_xlsx = args.output_xlsx or input_xlsx
        score_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet if args.sheet else None,
            max_ko_len=int(args.max_ko_len),
            max_zh_len=int(args.max_zh_len),
        )
        return

    if args.tag_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "Corpus(K2C)-2_20260526.xlsx")
        if args.output_xlsx:
            output_xlsx = args.output_xlsx
        else:
            base, ext = os.path.splitext(input_xlsx)
            output_xlsx = base + ".tagged.xlsx"
        keywords_file = args.keywords_file or os.path.join(args.root, "corpus_sources", "domain_keywords_ko.txt")
        tag_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet if args.sheet else None,
            keywords_path=keywords_file,
        )
        return

    if args.eval_set:
        translator_path = args.translator or os.path.join(args.root, "实时翻译测试_V3.0.1(greedy).py")
        eval_set(
            root=args.root,
            eval_path=args.eval_set,
            model_dir=model_dir,
            translator_path=translator_path,
            eval_out=args.eval_out,
            eval_sheet=args.eval_sheet,
            n=args.n,
            seed=args.seed,
            max_len=int(args.max_len),
        )
        return

    if args.clean_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "NEW-CORPUS-20260525.xlsx")
        output_xlsx = args.output_xlsx or _timestamped_clean_output_path(input_xlsx)
        clean_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet,
            min_hangul_ratio=float(args.min_hangul_ratio),
            min_hanzi_ratio=float(args.min_hanzi_ratio),
            min_punct_keep_ratio=float(args.min_punct_keep_ratio),
            min_alpha_keep_ratio=float(args.min_alpha_keep_ratio),
            min_digit_keep_ratio=float(args.min_digit_keep_ratio),
            max_ko_chars=int(args.clean_max_ko_chars),
            max_zh_chars=int(args.clean_max_zh_chars),
        )
        return

    if args.plugin_score:
        score_external_translators(
            life_eval_path=args.life_set,
            auto_eval_path=args.auto_set,
            baidu_path=args.baidu_plugin,
            kimi_path=args.kimi_plugin,
            baidu_output_xlsx=args.baidu_score_out,
            kimi_output_xlsx=args.kimi_score_out,
            n=(args.n if int(args.n) > 0 else None),
            seed=int(args.seed),
        )
        return

    if args.compare:
        v30_path = args.v30 or os.path.join(args.root, "实时翻译测试_V3.0(greedy).py")
        v301_path = args.v301 or os.path.join(args.root, "实时翻译测试_V3.0.1(greedy).py")
        compare_rt(
            root=args.root,
            corpus_xlsx=args.corpus,
            model_dir=model_dir,
            v30_path=v30_path,
            v301_path=v301_path,
            n=args.n,
            seed=args.seed,
        )
        return

    vocab_diagnose(root=args.root, corpus_xlsx=args.corpus, model_dir=model_dir, output_xlsx=args.vocab_diagnose_out)


if __name__ == "__main__":
    main()
